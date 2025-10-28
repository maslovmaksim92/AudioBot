from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Dict, Any, Optional
from datetime import datetime, timedelta, timezone
from pydantic import BaseModel
import logging
import asyncpg
import os
import io
import csv

logger = logging.getLogger(__name__)
router = APIRouter(tags=["finances"])

# Helper function для получения DB connection
async def get_db_connection():
    """Получить прямое соединение с БД для raw SQL"""
    db_url = os.environ.get('DATABASE_URL', '')
    
    if not db_url:
        raise HTTPException(status_code=500, detail="Database not configured")
    
    # Для asyncpg не нужно преобразовывать URL
    try:
        return await asyncpg.connect(db_url)
    except Exception as e:
        logger.error(f"Failed to connect to database: {e}")
        raise HTTPException(status_code=500, detail=f"Database connection failed: {str(e)}")

# Pydantic модели
class CashFlowItem(BaseModel):
    date: str
    income: float
    expense: float
    balance: float
    
class ProfitLossItem(BaseModel):
    period: str
    revenue: float
    expenses: float
    profit: float
    margin: float

class BalanceSheetItem(BaseModel):
    category: str
    assets: float
    liabilities: float
    equity: float

class ExpenseCategory(BaseModel):
    category: str
    amount: float
    percentage: float

class DebtItem(BaseModel):
    id: str
    creditor: str
    amount: float
    due_date: str
    status: str

class InventoryItem(BaseModel):
    id: str
    name: str
    quantity: int
    cost: float
    value: float

# API Endpoints

@router.get("/finances/cash-flow")
async def get_cash_flow(
    period: Optional[str] = "month"
):
    """
    Получить данные движения денег
    """
    try:
        conn = await get_db_connection()
        try:
            # Получаем данные по дням
            query = """
                SELECT 
                    DATE(date) as transaction_date,
                    SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) as income,
                    SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expense
                FROM financial_transactions
                GROUP BY DATE(date)
                ORDER BY DATE(date) DESC
                LIMIT 30
            """
            rows = await conn.fetch(query)
            
            cash_flow = []
            running_balance = 0
            
            # Сортируем от старых к новым для правильного расчета баланса
            for row in reversed(rows):
                income = float(row['income'])
                expense = float(row['expense'])
                balance = income - expense
                running_balance += balance
                
                cash_flow.append({
                    "date": row['transaction_date'].strftime("%Y-%m-%d"),
                    "income": income,
                    "expense": expense,
                    "balance": running_balance
                })
            
            # Возвращаем в обратном порядке (от новых к старым)
            cash_flow.reverse()
            
            if not cash_flow:
                # Если данных нет, возвращаем пустой результат
                return {
                    "cash_flow": [],
                    "summary": {
                        "total_income": 0,
                        "total_expense": 0,
                        "net_cash_flow": 0
                    }
                }
            
            return {
                "cash_flow": cash_flow,
                "summary": {
                    "total_income": sum(item["income"] for item in cash_flow),
                    "total_expense": sum(item["expense"] for item in cash_flow),
                    "net_cash_flow": sum(item["income"] for item in cash_flow) - sum(item["expense"] for item in cash_flow)
                }
            }
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching cash flow: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/finances/profit-loss")
async def get_profit_loss(
    period: Optional[str] = "year",
    company: Optional[str] = "ВАШ ДОМ ФАКТ"
):
    """
    Получить отчёт о прибылях и убытках
    Использует ручную выручку из monthly_revenue если она есть
    Параметры:
    - company: Фильтр по компании (по умолчанию "ВАШ ДОМ ФАКТ")
    """
    try:
        conn = await get_db_connection()
        try:
            # Если выбрана консолидация - используем специальную функцию
            if company == "ВАШ ДОМ модель":
                result = await get_consolidated_profit_loss(conn)
                return result
            # Проверяем существует ли таблица monthly_revenue
            table_exists = await conn.fetchval("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'monthly_revenue'
                )
            """)
            
            # Получаем ручную выручку если таблица существует
            manual_revenue = {}
            if table_exists:
                manual_rows = await conn.fetch("""
                    SELECT month, revenue
                    FROM monthly_revenue
                    WHERE company = $1
                """, company)
                manual_revenue = {row['month']: float(row['revenue']) for row in manual_rows}
            
            # Группируем данные по месяцам из даты
            query = """
                SELECT 
                    TO_CHAR(date, 'Month YYYY') as period,
                    TO_CHAR(date, 'YYYY-MM') as sort_key,
                    SUM(CASE WHEN type = 'income' THEN amount ELSE 0 END) as revenue,
                    SUM(CASE WHEN type = 'expense' THEN amount ELSE 0 END) as expenses
                FROM financial_transactions
                WHERE date IS NOT NULL AND EXTRACT(YEAR FROM date) = 2025 AND company = $1
                GROUP BY TO_CHAR(date, 'Month YYYY'), TO_CHAR(date, 'YYYY-MM')
                ORDER BY TO_CHAR(date, 'YYYY-MM')
            """
            rows = await conn.fetch(query, company)
            
            # Маппинг английских названий месяцев на русские
            month_map = {
                'January': 'Январь', 'February': 'Февраль', 'March': 'Март',
                'April': 'Апрель', 'May': 'Май', 'June': 'Июнь',
                'July': 'Июль', 'August': 'Август', 'September': 'Сентябрь',
                'October': 'Октябрь', 'November': 'Ноябрь', 'December': 'Декабрь'
            }
            
            profit_loss = []
            for row in rows:
                period_en = row['period'].strip()
                # Переводим на русский и нормализуем пробелы
                period_ru = period_en
                for eng, rus in month_map.items():
                    if eng in period_en:
                        period_ru = period_en.replace(eng, rus)
                        break
                # Убираем лишние пробелы между месяцем и годом
                period_ru = ' '.join(period_ru.split())
                
                # Используем ручную выручку если она есть, иначе из транзакций
                revenue = manual_revenue.get(period_ru, float(row['revenue']))
                expenses = float(row['expenses'])
                profit = revenue - expenses
                margin = (profit / revenue * 100) if revenue > 0 else 0
                
                profit_loss.append({
                    "period": period_ru,
                    "revenue": revenue,
                    "expenses": expenses,
                    "profit": profit,
                    "margin": round(margin, 2),
                    "manual_revenue": period_ru in manual_revenue
                })
            
            if not profit_loss:
                # Если данных нет, возвращаем пустой результат
                return {
                    "profit_loss": [],
                    "summary": {
                        "total_revenue": 0,
                        "total_expenses": 0,
                        "net_profit": 0,
                        "average_margin": 0
                    }
                }
            
            return {
                "profit_loss": profit_loss,
                "summary": {
                    "total_revenue": sum(item["revenue"] for item in profit_loss),
                    "total_expenses": sum(item["expenses"] for item in profit_loss),
                    "net_profit": sum(item["profit"] for item in profit_loss),
                    "average_margin": round(sum(item["margin"] for item in profit_loss) / len(profit_loss), 2) if profit_loss else 0
                }
            }
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching profit/loss: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/finances/balance-sheet")
async def get_balance_sheet():
    """
    Получить балансовый отчёт
    """
    try:
        # Mock данные
        balance = {
            "assets": {
                "current": {
                    "cash": 2500000,
                    "accounts_receivable": 1800000,
                    "inventory": 3200000,
                    "total": 7500000
                },
                "non_current": {
                    "property": 15000000,
                    "equipment": 8500000,
                    "vehicles": 5000000,
                    "total": 28500000
                },
                "total": 36000000
            },
            "liabilities": {
                "current": {
                    "accounts_payable": 1200000,
                    "short_term_debt": 800000,
                    "total": 2000000
                },
                "non_current": {
                    "long_term_debt": 10000000,
                    "total": 10000000
                },
                "total": 12000000
            },
            "equity": {
                "capital": 15000000,
                "retained_earnings": 9000000,
                "total": 24000000
            }
        }
        
        return balance
    except Exception as e:
        logger.error(f"Error fetching balance sheet: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/finances/expense-analysis")
async def get_expense_analysis(month: Optional[str] = None, company: Optional[str] = "ВАШ ДОМ ФАКТ"):
    """
    Анализ расходов по категориям
    Параметры:
    - month: Опциональный фильтр по месяцу (например, "Январь 2025")
    - company: Фильтр по компании (по умолчанию "ВАШ ДОМ ФАКТ")
    """
    try:
        conn = await get_db_connection()
        try:
            # Консолидированный расчет
            if company == "ВАШ ДОМ модель":
                return await get_consolidated_expenses(conn, month)
            
            # Получаем реальные данные из БД
            if month:
                query = """
                    SELECT category, SUM(amount) as total_amount
                    FROM financial_transactions
                    WHERE type = 'expense' AND project = $1 AND company = $2
                    GROUP BY category
                    ORDER BY total_amount DESC
                """
                rows = await conn.fetch(query, month, company)
            else:
                query = """
                    SELECT category, SUM(amount) as total_amount
                    FROM financial_transactions
                    WHERE type = 'expense' AND company = $1
                    GROUP BY category
                    ORDER BY total_amount DESC
                """
                rows = await conn.fetch(query, company)
            
            if not rows:
                # Если данных нет, возвращаем пустой результат
                return {
                    "expenses": [],
                    "total": 0,
                    "month": month
                }
            
            # Вычисляем общую сумму
            total = sum(float(row['total_amount']) for row in rows)
            
            # Формируем результат с процентами
            expenses = []
            for row in rows:
                amount = float(row['total_amount'])
                percentage = (amount / total * 100) if total > 0 else 0
                expenses.append({
                    "category": row['category'],
                    "amount": amount,
                    "percentage": round(percentage, 2)
                })
            
            return {
                "expenses": expenses,
                "total": total,
                "month": month
            }
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching expense analysis: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/finances/available-months")
async def get_available_months():
    """
    Получить список доступных месяцев с данными
    """
    try:
        conn = await get_db_connection()
        try:
            query = """
                SELECT DISTINCT project as month,
                       MIN(date) as start_date
                FROM financial_transactions
                WHERE project IS NOT NULL
                GROUP BY project
                ORDER BY start_date DESC
            """
            rows = await conn.fetch(query)
            
            months = [row['month'] for row in rows if row['month']]
            
            return {
                "months": months,
                "total": len(months)
            }
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching available months: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/finances/debts")
async def get_debts():
    """
    Получить список задолженностей из базы данных
    """
    try:
        conn = await get_db_connection()
        try:
            # Проверяем существует ли таблица debts
            table_exists = await conn.fetchval("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'debts'
                )
            """)
            
            if not table_exists:
                # Возвращаем пустой результат если таблицы нет
                return {
                    "debts": [],
                    "summary": {
                        "total": 0,
                        "overdue": 0,
                        "active": 0,
                        "count": 0
                    }
                }
            
            # Получаем все долги
            rows = await conn.fetch("""
                SELECT 
                    id, creditor, amount, due_date, status, type, description
                FROM debts
                ORDER BY due_date
            """)
            
            debts = []
            total_debt = 0
            overdue_debt = 0
            
            for row in rows:
                amount = float(row['amount'])
                total_debt += amount
                
                if row['status'] == 'overdue':
                    overdue_debt += amount
                
                debts.append({
                    "id": str(row['id']),
                    "creditor": row['creditor'],
                    "amount": amount,
                    "due_date": row['due_date'].strftime('%Y-%m-%d') if row['due_date'] else None,
                    "status": row['status'],
                    "type": row['type'],
                    "description": row['description']
                })
            
            return {
                "debts": debts,
                "summary": {
                    "total": total_debt,
                    "overdue": overdue_debt,
                    "active": total_debt - overdue_debt,
                    "count": len(debts)
                }
            }
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching debts: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/finances/inventory")
async def get_inventory():
    """
    Получить товарные запасы
    """
    try:
        # Mock данные
        inventory = [
            {
                "id": "inv-1",
                "name": "Моющие средства",
                "category": "Химия",
                "quantity": 500,
                "unit": "шт",
                "cost": 250,
                "value": 125000,
                "location": "Склад А"
            },
            {
                "id": "inv-2",
                "name": "Перчатки резиновые",
                "category": "Расходники",
                "quantity": 1000,
                "unit": "пар",
                "cost": 50,
                "value": 50000,
                "location": "Склад А"
            },
            {
                "id": "inv-3",
                "name": "Швабры",
                "category": "Инвентарь",
                "quantity": 150,
                "unit": "шт",
                "cost": 800,
                "value": 120000,
                "location": "Склад Б"
            },
            {
                "id": "inv-4",
                "name": "Ведра",
                "category": "Инвентарь",
                "quantity": 200,
                "unit": "шт",
                "cost": 300,
                "value": 60000,
                "location": "Склад Б"
            }
        ]
        
        total_value = sum(item["value"] for item in inventory)
        total_items = sum(item["quantity"] for item in inventory)
        
        return {
            "inventory": inventory,
            "summary": {
                "total_value": total_value,
                "total_items": total_items,
                "categories": len(set(item["category"] for item in inventory))
            }
        }
    except Exception as e:
        logger.error(f"Error fetching inventory: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/finances/dashboard")
async def get_finances_dashboard():
    """
    Получить сводку по всем финансовым показателям
    """
    try:
        # Комбинируем данные из всех эндпоинтов
        cash_flow = await get_cash_flow()
        profit_loss = await get_profit_loss()
        balance = await get_balance_sheet()
        expenses = await get_expense_analysis()
        debts = await get_debts()
        inventory = await get_inventory()
        
        return {
            "cash_flow": cash_flow["summary"],
            "profit_loss": profit_loss["summary"],
            "balance": {
                "total_assets": balance["assets"]["total"],
                "total_liabilities": balance["liabilities"]["total"],
                "total_equity": balance["equity"]["total"]
            },
            "expenses": expenses["total"],
            "debts": debts["summary"],
            "inventory": inventory["summary"]
        }
    except Exception as e:
        logger.error(f"Error getting dashboard: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/finances/export-expenses")
async def export_expenses(year: int = 2025):
    """
    Экспорт расходов в CSV формат помесячно
    """
    try:
        conn = await get_db_connection()
        try:
            # Получаем расходы по месяцам с детализацией по категориям
            query = """
                SELECT 
                    TO_CHAR(date, 'YYYY-MM') as month,
                    TO_CHAR(date, 'Month YYYY') as month_name,
                    category,
                    SUM(amount) as total_amount,
                    COUNT(*) as transactions_count
                FROM financial_transactions
                WHERE type = 'expense' AND EXTRACT(YEAR FROM date) = $1
                GROUP BY TO_CHAR(date, 'YYYY-MM'), TO_CHAR(date, 'Month YYYY'), category
                ORDER BY month, category
            """
            rows = await conn.fetch(query, year)
            
            # Создаем CSV в памяти
            output = io.StringIO()
            writer = csv.writer(output)
            
            # Заголовки
            writer.writerow(['Месяц', 'Категория', 'Сумма расходов (₽)', 'Количество транзакций'])
            
            # Данные
            for row in rows:
                writer.writerow([
                    row['month_name'].strip(),
                    row['category'],
                    float(row['total_amount']),
                    row['transactions_count']
                ])
            
            # Добавляем итоги по месяцам
            writer.writerow([])
            writer.writerow(['ИТОГО ПО МЕСЯЦАМ:'])
            writer.writerow(['Месяц', 'Всего расходов (₽)'])
            
            monthly_query = """
                SELECT 
                    TO_CHAR(date, 'Month YYYY') as month_name,
                    SUM(amount) as total_amount
                FROM financial_transactions
                WHERE type = 'expense' AND EXTRACT(YEAR FROM date) = $1
                GROUP BY TO_CHAR(date, 'YYYY-MM'), TO_CHAR(date, 'Month YYYY')
                ORDER BY TO_CHAR(date, 'YYYY-MM')
            """
            monthly_rows = await conn.fetch(monthly_query, year)
            
            for row in monthly_rows:
                writer.writerow([
                    row['month_name'].strip(),
                    float(row['total_amount'])
                ])
            
            # Возвращаем CSV файл
            output.seek(0)
            
            return StreamingResponse(
                io.BytesIO(output.getvalue().encode('utf-8-sig')),  # utf-8-sig для правильного отображения в Excel
                media_type='text/csv',
                headers={
                    'Content-Disposition': f'attachment; filename=expenses_{year}.csv'
                }
            )
            
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error exporting expenses: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/finances/expense-details")
async def get_expense_details(month: str, company: Optional[str] = "ООО ВАШ ДОМ"):
    """
    Получить детальные транзакции расходов для конкретного месяца
    Параметры:
    - month: Название месяца (например, "Июль 2025")
    - company: Фильтр по компании (по умолчанию "ООО ВАШ ДОМ")
    """
    try:
        conn = await get_db_connection()
        try:
            # Получаем все транзакции расходов для указанного месяца
            query = """
                SELECT 
                    id,
                    date,
                    amount,
                    category,
                    description,
                    payment_method,
                    counterparty
                FROM financial_transactions
                WHERE type = 'expense' AND project = $1 AND company = $2
                ORDER BY date DESC, category
            """
            rows = await conn.fetch(query, month, company)
            
            if not rows:
                return {
                    "transactions": [],
                    "total": 0,
                    "month": month,
                    "count": 0
                }
            
            # Формируем список транзакций
            transactions = []
            total = 0
            
            for row in rows:
                amount = float(row['amount'])
                total += amount
                
                transactions.append({
                    "id": row['id'],
                    "date": row['date'].strftime('%d.%m.%Y') if row['date'] else "",
                    "category": row['category'],
                    "amount": amount,
                    "description": row['description'] or "",
                    "payment_method": row['payment_method'] or "",
                    "counterparty": row['counterparty'] or ""
                })
            
            return {
                "transactions": transactions,
                "total": total,
                "month": month,
                "count": len(transactions)
            }
            
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching expense details: {e}")


@router.get("/finances/revenue-analysis")
async def get_revenue_analysis(month: Optional[str] = None, company: Optional[str] = "ВАШ ДОМ ФАКТ"):
    """
    Анализ выручки по категориям
    Параметры:
    - month: Опциональный фильтр по месяцу (например, "Январь 2025")
    - company: Фильтр по компании (по умолчанию "ВАШ ДОМ ФАКТ")
    """
    try:
        conn = await get_db_connection()
        try:
            # Консолидированный расчет - выручка ООО ВАШ ДОМ минус "Швеи" и "Аутсорсинг"
            if company == "ВАШ ДОМ модель":
                return await get_consolidated_revenue(conn, month)
            
            # Получаем реальные данные из БД
            if month:
                query = """
                    SELECT category, SUM(amount) as total_amount
                    FROM financial_transactions
                    WHERE type = 'income' AND project = $1 AND company = $2
                    GROUP BY category
                    ORDER BY total_amount DESC
                """
                rows = await conn.fetch(query, month, company)
            else:
                query = """
                    SELECT category, SUM(amount) as total_amount
                    FROM financial_transactions
                    WHERE type = 'income' AND company = $1
                    GROUP BY category
                    ORDER BY total_amount DESC
                """
                rows = await conn.fetch(query, company)
            
            if not rows:
                # Если данных нет, возвращаем пустой результат
                return {
                    "revenue": [],
                    "total": 0,
                    "month": month
                }
            
            # Вычисляем общую сумму
            total = sum(float(row['total_amount']) for row in rows)
            
            # Формируем результат с процентами
            revenue = []
            for row in rows:
                amount = float(row['total_amount'])
                percentage = (amount / total * 100) if total > 0 else 0
                revenue.append({
                    "category": row['category'],
                    "amount": amount,
                    "percentage": round(percentage, 2)
                })
            
            return {
                "revenue": revenue,
                "total": total,
                "month": month
            }
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching revenue analysis: {e}")
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/finances/revenue-details")
async def get_revenue_details(month: Optional[str] = None, company: Optional[str] = "ВАШ ДОМ ФАКТ"):
    """
    Получить детальные транзакции доходов
    Параметры:
    - month: Опциональный фильтр по месяцу (например, "Январь 2025")
    - company: Фильтр по компании (по умолчанию "ВАШ ДОМ ФАКТ")
    """
    try:
        conn = await get_db_connection()
        try:
            # Получаем все транзакции доходов
            if month:
                query = """
                    SELECT 
                        id, date, amount, category, description, counterparty
                    FROM financial_transactions
                    WHERE type = 'income' AND project = $1 AND company = $2
                    ORDER BY date DESC
                """
                rows = await conn.fetch(query, month, company)
            else:
                query = """
                    SELECT 
                        id, date, amount, category, description, counterparty
                    FROM financial_transactions
                    WHERE type = 'income' AND company = $1
                    ORDER BY date DESC
                """
                rows = await conn.fetch(query, company)
            
            if not rows:
                return {
                    "transactions": [],
                    "total": 0,
                    "month": month,
                    "count": 0
                }
            
            # Формируем список транзакций
            transactions = []
            total = 0
            
            for row in rows:
                amount = float(row['amount'])
                total += amount
                
                transactions.append({
                    "id": str(row['id']),
                    "date": row['date'].strftime('%d.%m.%Y') if row['date'] else "",
                    "counterparty": row['counterparty'] or "—",
                    "amount": amount,
                    "description": row['description'] or "",
                    "category": row['category']
                })
            
            return {
                "transactions": transactions,
                "total": total,
                "month": month,
                "count": len(transactions)
            }
            
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error fetching revenue details: {e}")
        raise HTTPException(status_code=500, detail=str(e))


async def get_consolidated_revenue(conn, month: Optional[str] = None):
    """
    Консолидированная выручка для "ВАШ ДОМ модель"
    Используется ручная выручка из monthly_revenue
    """
    # Для консолидированной модели используем ручную выручку из monthly_revenue
    if month:
        query = """
            SELECT month, revenue
            FROM monthly_revenue
            WHERE company = 'ВАШ ДОМ модель' AND month = $1
        """
        rows = await conn.fetch(query, month)
    else:
        query = """
            SELECT month, revenue
            FROM monthly_revenue
            WHERE company = 'ВАШ ДОМ модель'
            ORDER BY month
        """
        rows = await conn.fetch(query)
    
    if not rows:
        return {
            "revenue": [],
            "total": 0,
            "month": month
        }
    
    # Вычисляем общую сумму
    total = sum(float(row['revenue']) for row in rows)
    
    # Формируем результат
    revenue = [{
        "category": "Консолидированная выручка",
        "amount": total,
        "percentage": 100.0
    }]
    
    return {
        "revenue": revenue,
        "total": total,
        "month": month
    }


async def get_consolidated_profit_loss(conn):
    """
    Консолидированный расчет для "ВАШ ДОМ модель"
    
    Логика:
    - Выручка: из monthly_revenue для "ВАШ ДОМ модель" (ручная)
    - Расходы:
      * Большинство категорий ("так же"): только ВАШ ДОМ ФАКТ
      * Зарплата: ВАШ ДОМ ФАКТ - УФИЦ модель (ТОЛЬКО Зарплата, без ФОТ) помесячно
      * Налоги: 5% от выручки каждого месяца
      * Аутсорсинг персонала: из транзакций ВАШ ДОМ модель
      * Исключить: Кредиты, Швеи, Юридические услуги, Продукты питания
    """
    # Получаем ручную выручку для консолидированной модели
    revenue_rows = await conn.fetch("""
        SELECT month, revenue
        FROM monthly_revenue
        WHERE company = 'ВАШ ДОМ модель'
        ORDER BY month
    """)
    
    revenue_by_month = {row['month']: float(row['revenue']) for row in revenue_rows}
    
    # Получаем расходы ВАШ ДОМ ФАКТ по месяцам и категориям
    vasdom_expenses = await conn.fetch("""
        SELECT 
            project as month,
            category,
            SUM(amount) as amount
        FROM financial_transactions
        WHERE type = 'expense' AND company = 'ВАШ ДОМ ФАКТ'
        GROUP BY project, category
    """)
    
    # Получаем расходы УФИЦ модель по месяцам - ТОЛЬКО Зарплата (без ФОТ)
    ufic_expenses = await conn.fetch("""
        SELECT 
            project as month,
            category,
            SUM(amount) as amount
        FROM financial_transactions
        WHERE type = 'expense' AND company = 'УФИЦ модель'
          AND category = 'Зарплата'
        GROUP BY project, category
    """)
    
    # Получаем Аутсорсинг персонала из транзакций ВАШ ДОМ модель
    outsourcing_expenses = await conn.fetch("""
        SELECT 
            project as month,
            SUM(amount) as amount
        FROM financial_transactions
        WHERE type = 'expense' AND company = 'ВАШ ДОМ модель'
          AND category = 'Аутсорсинг персонала'
        GROUP BY project
    """)
    
    # Группируем расходы ВАШ ДОМ ФАКТ по месяцам
    vasdom_by_month = {}
    for row in vasdom_expenses:
        month = row['month']
        category = row['category']
        amount = float(row['amount'])
        
        if month not in vasdom_by_month:
            vasdom_by_month[month] = {}
        vasdom_by_month[month][category] = amount
    
    # Группируем расходы УФИЦ по месяцам (только Зарплата)
    ufic_salary_by_month = {}
    for row in ufic_expenses:
        month = row['month']
        amount = float(row['amount'])
        ufic_salary_by_month[month] = amount
    
    # Группируем Аутсорсинг персонала по месяцам
    outsourcing_by_month = {}
    for row in outsourcing_expenses:
        month = row['month']
        amount = float(row['amount'])
        outsourcing_by_month[month] = amount
    
    # Получаем все уникальные месяцы
    all_months_set = set(list(revenue_by_month.keys()) + list(vasdom_by_month.keys()) + list(outsourcing_by_month.keys()))
    
    # Правильная сортировка месяцев
    month_order = {
        'Январь 2025': 1, 'Февраль 2025': 2, 'Март 2025': 3,
        'Апрель 2025': 4, 'Май 2025': 5, 'Июнь 2025': 6,
        'Июль 2025': 7, 'Август 2025': 8, 'Сентябрь 2025': 9,
        'Октябрь 2025': 10, 'Ноябрь 2025': 11, 'Декабрь 2025': 12
    }
    all_months = sorted(all_months_set, key=lambda x: month_order.get(x, 99))
    
    # Вычисляем консолидированные данные
    profit_loss = []
    total_revenue = 0
    total_expenses = 0
    
    for month in all_months:
        revenue = revenue_by_month.get(month, 0)
        
        # Вычисляем расходы
        vasdom_exp = vasdom_by_month.get(month, {})
        
        month_expenses = 0
        
        for category, amount in vasdom_exp.items():
            # Исключаем категории
            if category in ['Кредиты', 'Швеи', 'Аутсорсинг персонала с Ю/ЦЛ', 'Юридические услуги', 'Продукты питания', 'Налоги']:
                continue
            
            # Специальная логика для Зарплаты: ВАШ ДОМ ФАКТ - УФИЦ Зарплата (без ФОТ)
            if category == 'Зарплата':
                ufic_salary = ufic_salary_by_month.get(month, 0)
                consolidated_salary = amount - ufic_salary
                month_expenses += max(0, consolidated_salary)
            else:
                # Для остальных категорий: только ВАШ ДОМ ФАКТ ("так же")
                month_expenses += amount
        
        # Добавляем Аутсорсинг персонала из транзакций ВАШ ДОМ модель
        outsourcing = outsourcing_by_month.get(month, 0)
        month_expenses += outsourcing
        
        # Добавляем Налоги: 5% от выручки
        taxes = revenue * 0.05
        month_expenses += taxes
        
        profit = revenue - month_expenses
        margin = (profit / revenue * 100) if revenue > 0 else 0
        
        profit_loss.append({
            "period": month,
            "revenue": revenue,
            "expenses": month_expenses,
            "profit": profit,
            "margin": round(margin, 2)
        })
        
        total_revenue += revenue
        total_expenses += month_expenses
    
    total_profit = total_revenue - total_expenses
    total_margin = (total_profit / total_revenue * 100) if total_revenue > 0 else 0
    
    return {
        "profit_loss": profit_loss,
        "summary": {
            "total_revenue": total_revenue,
            "total_expenses": total_expenses,
            "total_profit": total_profit,
            "margin": round(total_margin, 2)
        }
    }


async def get_consolidated_expenses(conn, month: Optional[str] = None):
    """
    Консолидированные расходы для "ВАШ ДОМ модель"
    
    Логика:
    - Большинство категорий ("так же"): только ВАШ ДОМ ФАКТ
    - Зарплата: ВАШ ДОМ ФАКТ - УФИЦ Зарплата (ТОЛЬКО Зарплата, без ФОТ) помесячно
    - Налоги: 5% от выручки каждого месяца
    - Аутсорсинг персонала: из транзакций ВАШ ДОМ модель
    - Исключить: Кредиты, Швеи, Юридические услуги, Продукты питания
    """
    # Получаем выручку для расчета налогов
    if month:
        revenue_query = """
            SELECT month, revenue
            FROM monthly_revenue
            WHERE company = 'ВАШ ДОМ модель' AND month = $1
        """
        revenue_rows = await conn.fetch(revenue_query, month)
    else:
        revenue_query = """
            SELECT month, revenue
            FROM monthly_revenue
            WHERE company = 'ВАШ ДОМ модель'
        """
        revenue_rows = await conn.fetch(revenue_query)
    
    revenue_by_month = {row['month']: float(row['revenue']) for row in revenue_rows}
    
    # Получаем Аутсорсинг персонала из транзакций ВАШ ДОМ модель
    if month:
        outsourcing_query = """
            SELECT SUM(amount) as total_amount
            FROM financial_transactions
            WHERE type = 'expense' AND company = 'ВАШ ДОМ модель' 
              AND category = 'Аутсорсинг персонала' AND project = $1
        """
        outsourcing_row = await conn.fetchrow(outsourcing_query, month)
    else:
        outsourcing_query = """
            SELECT SUM(amount) as total_amount
            FROM financial_transactions
            WHERE type = 'expense' AND company = 'ВАШ ДОМ модель'
              AND category = 'Аутсорсинг персонала'
        """
        outsourcing_row = await conn.fetchrow(outsourcing_query)
    
    outsourcing_amount = float(outsourcing_row['total_amount']) if outsourcing_row and outsourcing_row['total_amount'] else 0
    
    # Получаем расходы ВАШ ДОМ ФАКТ (исключая определенные категории)
    exclude_categories = ['Кредиты', 'Швеи', 'Юридические услуги', 'Продукты питания']
    
    if month:
        vasdom_query = """
            SELECT category, SUM(amount) as total_amount
            FROM financial_transactions
            WHERE type = 'expense' AND company = 'ВАШ ДОМ ФАКТ' 
              AND project = $1
              AND category NOT IN ('Кредиты', 'Швеи', 'Юридические услуги', 'Продукты питания')
            GROUP BY category
        """
        vasdom_rows = await conn.fetch(vasdom_query, month)
    else:
        vasdom_query = """
            SELECT category, SUM(amount) as total_amount
            FROM financial_transactions
            WHERE type = 'expense' AND company = 'ВАШ ДОМ ФАКТ'
              AND category NOT IN ('Кредиты', 'Швеи', 'Юридические услуги', 'Продукты питания')
            GROUP BY category
        """
        vasdom_rows = await conn.fetch(vasdom_query)
    
    # Получаем зарплату УФИЦ для вычитания
    if month:
        ufic_salary_query = """
            SELECT SUM(amount) as total_amount
            FROM financial_transactions
            WHERE type = 'expense' AND company = 'УФИЦ модель' 
              AND category = 'Зарплата' AND project = $1
        """
        ufic_salary_row = await conn.fetchrow(ufic_salary_query, month)
    else:
        ufic_salary_query = """
            SELECT SUM(amount) as total_amount
            FROM financial_transactions
            WHERE type = 'expense' AND company = 'УФИЦ модель'
              AND category = 'Зарплата'
        """
        ufic_salary_row = await conn.fetchrow(ufic_salary_query)
    
    ufic_salary = float(ufic_salary_row['total_amount']) if ufic_salary_row and ufic_salary_row['total_amount'] else 0
    
    # Формируем результат
    expenses_dict = {}
    
    # Добавляем расходы ВАШ ДОМ ФАКТ
    for row in vasdom_rows:
        category = row['category']
        amount = float(row['total_amount'])
        
        # Зарплата: вычитаем УФИЦ зарплату
        if category == 'Зарплата':
            amount = amount - ufic_salary
            if amount < 0:
                amount = 0
        
        expenses_dict[category] = expenses_dict.get(category, 0) + amount
    
    # Добавляем аутсорсинг персонала
    if outsourcing_amount > 0:
        expenses_dict['Аутсорсинг персонала'] = outsourcing_amount
    
    # Добавляем налоги (5% от выручки)
    total_revenue = sum(revenue_by_month.values())
    if total_revenue > 0:
        taxes = total_revenue * 0.05
        expenses_dict['Налоги'] = taxes
    
    # Вычисляем общую сумму и проценты
    total = sum(expenses_dict.values())
    
    expenses = []
    for category, amount in sorted(expenses_dict.items(), key=lambda x: x[1], reverse=True):
        percentage = (amount / total * 100) if total > 0 else 0
        expenses.append({
            "category": category,
            "amount": amount,
            "percentage": round(percentage, 2)
        })
    
    return {
        "expenses": expenses,
        "total": total,
        "month": month
    }


@router.get("/finances/forecast")
async def get_forecast(
    company: Optional[str] = "ВАШ ДОМ ФАКТ",
    scenario: Optional[str] = "realistic"
):
    """Get forecast for 2026-2030 with three scenarios"""
    try:
        conn = await get_db_connection()
        try:
            # Коэффициенты роста для разных сценариев
            scenario_coefficients = {
                "pessimistic": {
                    "revenue_multiplier": 0.7,  # Снижение роста выручки на 30%
                    "expense_multiplier": 1.3,  # Увеличение роста расходов на 30%
                    "description": "Консервативный прогноз: снижение темпов роста выручки, увеличение расходов"
                },
                "realistic": {
                    "revenue_multiplier": 1.0,  # Базовый рост
                    "expense_multiplier": 1.0,  # Базовый рост
                    "description": "Базовый прогноз на основе текущих трендов 2025 года"
                },
                "optimistic": {
                    "revenue_multiplier": 1.4,  # Увеличение роста выручки на 40%
                    "expense_multiplier": 0.8,  # Снижение роста расходов на 20%
                    "description": "Оптимистичный прогноз: ускоренный рост выручки, оптимизация расходов"
                }
            }
            
            scenario_config = scenario_coefficients.get(scenario, scenario_coefficients["realistic"])
            
            # Для УФИЦ модель используем точные данные из Excel файла "Модель УФИЦ.xlsx"
            if company == "УФИЦ модель":
                # Данные из Excel модели по сценариям (2025 и 2026 годы) с детализацией
                ufic_excel_data = {
                    "pessimistic": {
                        "revenue_2025": 38467135.00,
                        "expenses_2025": 27195041.00,
                        "revenue_2026": 49454100.00,
                        "expenses_2026": 34148893.00,
                        "description": "Консервативный прогноз. Швеи: 10, Уборщицы: 45, Аутсорсинг: 5. Индексация 4.8% ежегодно с 2027 года.",
                        "detailed_description": {
                            "summary": "Консервативный сценарий с базовой загрузкой и фокусом на уборку",
                            "revenue_factors": [
                                "Стабильная работа с текущими клиентами управляющих компаний",
                                "Минимальное повышение цен на услуги (индексация 4.8%)",
                                "Фокус на надежности и качества услуг уборки"
                            ],
                            "expense_factors": [
                                "ФОТ (зарплата): 39,200 ₽/сотрудник",
                                "ФОТ управляющий персонал: 5,200 ₽/сотрудник",
                                "Обслуживание помещений: 3,029 ₽/сотрудник"
                            ],
                            "unit_economics": {
                                "revenue_per_employee": 68686,
                                "expense_per_employee": 47429,
                                "profit_per_employee": 21257,
                                "total_employees": 60
                            },
                            "advantages": [
                                "Максимальная загрузка уборщицами (45 чел) обеспечивает стабильную выручку от ключевого заказчика",
                                "Простая операционная модель с фокусом на основном направлении",
                                "Минимальные риски за счет небольшого штата швей и аутсорсинга"
                            ],
                            "disadvantages": [
                                "Высокая зависимость от одного направления услуг (уборка)",
                                "Малый штат квалифицированных сотрудников ограничивает потенциал роста",
                                "Уязвимость к потере ключевого клиента клининговой компании"
                            ]
                        },
                        "revenue_breakdown_2026": {
                            "sewing": 10696500.00,
                            "cleaning": 34347600.00,
                            "outsourcing": 4410000.00
                        },
                        "expense_breakdown_2026": {
                            "labor": 34148893.00
                        },
                        "staff_counts": {
                            "sewing": 10,
                            "cleaning": 45,
                            "outsourcing": 5
                        }
                    },
                    "realistic": {
                        "revenue_2025": 38467135.00,
                        "expenses_2025": 27195042.00,
                        "revenue_2026": 55171343.00,
                        "expenses_2026": 36994634.00,
                        "description": "Реалистичный прогноз. Швеи: 15, Уборщицы: 30, Аутсорсинг: 20. Индексация 4.8% ежегодно с 2027 года.",
                        "detailed_description": {
                            "summary": "Реалистичный сценарий с оптимальной диверсификацией услуг",
                            "revenue_factors": [
                                "Оптимальная загрузка рабочих мест по всем направлениям",
                                "Расширение портфеля услуг (швеи, уборка, аутсорсинг)",
                                "Стабильный рост клиентской базы управляющих компаний"
                            ],
                            "expense_factors": [
                                "ФОТ (зарплата): 39,200 ₽/сотрудник",
                                "ФОТ управляющий персонал: 4,800 ₽/сотрудник",
                                "Обслуживание помещений: 3,429 ₽/сотрудник"
                            ],
                            "unit_economics": {
                                "revenue_per_employee": 70732,
                                "expense_per_employee": 47429,
                                "profit_per_employee": 23303,
                                "total_employees": 65
                            },
                            "advantages": [
                                "Сбалансированная диверсификация услуг (швеи 15, уборщицы 30, аутсорсинг 20)",
                                "Максимальная выручка на сотрудника (70,732₽) среди всех сценариев",
                                "Оптимальный баланс между ростом и стабильностью"
                            ],
                            "disadvantages": [
                                "Средняя загрузка уборщицами (30 чел) снижает потенциал выручки",
                                "Необходимость развития компетенций в трех направлениях одновременно",
                                "Зависимость от успешного развития направления аутсорсинга"
                            ]
                        },
                        "revenue_breakdown_2026": {
                            "sewing": 12033563.00,
                            "cleaning": 24615780.00,
                            "outsourcing": 18522000.00
                        },
                        "expense_breakdown_2026": {
                            "labor": 36994634.00
                        },
                        "staff_counts": {
                            "sewing": 15,
                            "cleaning": 30,
                            "outsourcing": 20
                        }
                    },
                    "optimistic": {
                        "revenue_2025": 38687635.00,
                        "expenses_2025": 27396013.00,
                        "revenue_2026": 59856630.00,
                        "expenses_2026": 39840376.00,
                        "description": "Оптимистичный прогноз. Швеи: 20, Уборщицы: 30, Аутсорсинг: 30. Индексация 6.9% ежегодно с 2027 года.",
                        "detailed_description": {
                            "summary": "Оптимистичный сценарий с максимальным развитием аутсорсинга",
                            "revenue_factors": [
                                "Активное развитие аутсорсингового направления (30 чел)",
                                "Привлечение крупных корпоративных контрактов",
                                "Повышенная индексация 6.9% ежегодно"
                            ],
                            "expense_factors": [
                                "ФОТ (зарплата): 39,200 ₽/сотрудник",
                                "ФОТ управляющий персонал: 4,460 ₽/сотрудник",
                                "Обслуживание помещений: 3,769 ₽/сотрудник"
                            ],
                            "unit_economics": {
                                "revenue_per_employee": 71258,
                                "expense_per_employee": 47429,
                                "profit_per_employee": 23829,
                                "total_employees": 70
                            },
                            "advantages": [
                                "Максимальное развитие аутсорсинга (30 чел) - наиболее прибыльное направление",
                                "Крупнейший штат (70 чел) позволяет обслуживать больше клиентов",
                                "Наивысшая прибыль на сотрудника (23,829₽)"
                            ],
                            "disadvantages": [
                                "Минимальная загрузка уборщицами (30 чел) среди всех сценариев",
                                "Сильная зависимость от успеха аутсорсингового направления",
                                "Необходимость значительных инвестиций в обучение и развитие"
                            ]
                        },
                        "revenue_breakdown_2026": {
                            "sewing": 16044750.00,
                            "cleaning": 16028880.00,
                            "outsourcing": 27783000.00
                        },
                        "expense_breakdown_2026": {
                            "labor": 39840376.00
                        },
                        "staff_counts": {
                            "sewing": 20,
                            "cleaning": 30,
                            "outsourcing": 30
                        }
                    }
                }
                
                # Генерируем прогноз с годовой индексацией 6%
                indexation_rate = 1.06
                
                ufic_data = {
                    "pessimistic": {
                        "years": [],
                        "base_revenue": ufic_excel_data["pessimistic"]["revenue_2025"],
                        "base_expenses": ufic_excel_data["pessimistic"]["expenses_2025"],
                        "description": ufic_excel_data["pessimistic"]["description"]
                    },
                    "realistic": {
                        "years": [],
                        "base_revenue": ufic_excel_data["realistic"]["revenue_2025"],
                        "base_expenses": ufic_excel_data["realistic"]["expenses_2025"],
                        "description": ufic_excel_data["realistic"]["description"]
                    },
                    "optimistic": {
                        "years": [],
                        "base_revenue": ufic_excel_data["optimistic"]["revenue_2025"],
                        "base_expenses": ufic_excel_data["optimistic"]["expenses_2025"],
                        "description": ufic_excel_data["optimistic"]["description"]
                    }
                }
                
                # Получаем детализацию расходов УФИЦ за 2025 из БД
                ufic_expense_categories = await conn.fetch("""
                    SELECT 
                        category,
                        SUM(amount) as total_amount
                    FROM financial_transactions
                    WHERE type = 'expense' AND company = 'УФИЦ модель'
                    GROUP BY category
                    ORDER BY total_amount DESC
                """)
                
                # Формируем структуру расходов 2025
                ufic_expenses_2025_breakdown = {}
                
                # Переменные для объединения категорий
                repair_training_total = 0
                
                for row in ufic_expense_categories:
                    category = row['category']
                    amount = float(row['total_amount'])
                    category_name = category.lower().replace(' ', '_').replace('-', '_')
                    
                    # Пропускаем НДФЛ - не включаем в структуру
                    if 'ндфл' in category_name or 'ndfl' in category_name:
                        continue
                    
                    # Объединяем категории в "Текущий ремонт, обучение"
                    if any(keyword in category.lower() for keyword in ['газпром', 'первый газовый', 'водоканал', 'крэо', 'вдпо', 'прикамский институт']):
                        repair_training_total += amount
                    else:
                        ufic_expenses_2025_breakdown[category_name] = amount
                
                # Добавляем объединенную категорию
                if repair_training_total > 0:
                    ufic_expenses_2025_breakdown['текущий_ремонт_обучение'] = repair_training_total
                
                total_ufic_expenses_2025 = sum(ufic_expenses_2025_breakdown.values())
                
                # Заполняем данные для всех сценариев
                for scen_name in ["pessimistic", "realistic", "optimistic"]:
                    scen_excel = ufic_excel_data[scen_name]
                    
                    # Индексация: 4.8% для пессимистичного и реалистичного, 6.9% для оптимистичного
                    indexation_rate = 1.069 if scen_name == "optimistic" else 1.048
                    
                    # 2026 год - берем из Excel с детализацией
                    revenue_2026 = scen_excel["revenue_2026"]
                    expenses_2026 = scen_excel["expenses_2026"]
                    profit_2026 = revenue_2026 - expenses_2026
                    
                    revenue_breakdown_2026 = scen_excel["revenue_breakdown_2026"]
                    
                    # Детализация расходов 2026 - пропорционально от 2025
                    expense_breakdown_2026 = {}
                    if total_ufic_expenses_2025 > 0:
                        expenses_multiplier = expenses_2026 / total_ufic_expenses_2025
                        for category, amount_2025 in ufic_expenses_2025_breakdown.items():
                            expense_breakdown_2026[category] = round(amount_2025 * expenses_multiplier, 2)
                    else:
                        # Если нет данных, используем только labor
                        expense_breakdown_2026 = {"labor": expenses_2026}
                    
                    ufic_data[scen_name]["years"].append({
                        "year": 2026,
                        "revenue": round(revenue_2026, 2),
                        "expenses": round(expenses_2026, 2),
                        "profit": round(profit_2026, 2),
                        "cleaners": 0,  # Для совместимости
                        "revenue_breakdown": {
                            "sewing": round(revenue_breakdown_2026["sewing"], 2),
                            "cleaning": round(revenue_breakdown_2026["cleaning"], 2),
                            "outsourcing": round(revenue_breakdown_2026["outsourcing"], 2)
                        },
                        "expense_breakdown": {
                            category: round(amount, 2) for category, amount in expense_breakdown_2026.items()
                        }
                    })
                    
                    # 2027-2030 - применяем индексацию к данным 2026 (включая детализацию)
                    for year in range(2027, 2031):
                        years_from_2026 = year - 2026
                        indexed_revenue = revenue_2026 * (indexation_rate ** years_from_2026)
                        indexed_expenses = expenses_2026 * (indexation_rate ** years_from_2026)
                        indexed_profit = indexed_revenue - indexed_expenses
                        
                        # Индексируем детализацию
                        indexed_revenue_breakdown = {
                            "sewing": round(revenue_breakdown_2026["sewing"] * (indexation_rate ** years_from_2026), 2),
                            "cleaning": round(revenue_breakdown_2026["cleaning"] * (indexation_rate ** years_from_2026), 2),
                            "outsourcing": round(revenue_breakdown_2026["outsourcing"] * (indexation_rate ** years_from_2026), 2)
                        }
                        
                        indexed_expense_breakdown = {
                            category: round(amount * (indexation_rate ** years_from_2026), 2) 
                            for category, amount in expense_breakdown_2026.items()
                        }
                        
                        ufic_data[scen_name]["years"].append({
                            "year": year,
                            "revenue": round(indexed_revenue, 2),
                            "expenses": round(indexed_expenses, 2),
                            "profit": round(indexed_profit, 2),
                            "cleaners": 0,  # Для совместимости
                            "revenue_breakdown": indexed_revenue_breakdown,
                            "expense_breakdown": indexed_expense_breakdown
                        })
                    
                    ufic_data[scen_name]["base_revenue"] = scen_excel["revenue_2025"]
                    ufic_data[scen_name]["base_expenses"] = scen_excel["expenses_2025"]
                
                ufic_scenario = ufic_data.get(scenario, ufic_data["realistic"])
                
                # Используем данные из Excel
                forecast = []
                for year_data in ufic_scenario["years"]:
                    margin = (year_data["profit"] / year_data["revenue"] * 100) if year_data["revenue"] > 0 else 0
                    forecast_item = {
                        "year": year_data["year"],
                        "revenue": year_data["revenue"],
                        "expenses": year_data["expenses"],
                        "profit": year_data["profit"],
                        "margin": round(margin, 2),
                        "cleaners_count": year_data["cleaners"]  # Добавляем количество уборщиц
                    }
                    # Добавляем детализацию доходов и расходов если есть
                    if "revenue_breakdown" in year_data:
                        forecast_item["revenue_breakdown"] = year_data["revenue_breakdown"]
                    if "expense_breakdown" in year_data:
                        forecast_item["expense_breakdown"] = year_data["expense_breakdown"]
                    forecast.append(forecast_item)
                
                # Базовые данные 2025
                base_revenue = ufic_scenario["base_revenue"]
                base_expenses = ufic_scenario["base_expenses"]
                base_profit = base_revenue - base_expenses
                base_margin = (base_profit / base_revenue * 100) if base_revenue > 0 else 0
                
                # Расчеты для инвестора
                total_forecast_profit = sum(f["profit"] for f in forecast)
                average_annual_profit = total_forecast_profit / len(forecast)
                average_margin = sum(f["margin"] for f in forecast) / len(forecast)
                investment_amount = 40000000  # 40 млн рублей для всех моделей
                roi_5_years = (total_forecast_profit / investment_amount * 100) if investment_amount > 0 else 0
                
                # Срок окупаемости
                cumulative_profit = 0
                payback_period = None
                for i, f in enumerate(forecast):
                    cumulative_profit += f["profit"]
                    if cumulative_profit >= investment_amount and payback_period is None:
                        payback_period = i + 1
                
                if payback_period is None:
                    payback_period = "> 5 лет"
                
                # Рост выручки и расходов (из первого и последнего года)
                revenue_growth = ((forecast[-1]["revenue"] / forecast[0]["revenue"]) ** (1/5) - 1) * 100
                expense_growth = ((forecast[-1]["expenses"] / forecast[0]["expenses"]) ** (1/5) - 1) * 100
                
                return {
                    "company": company,
                    "scenario": scenario,
                    "base_year": 2025,
                    "base_data": {
                        "revenue": round(base_revenue, 2),
                        "expenses": round(base_expenses, 2),
                        "profit": round(base_profit, 2),
                        "margin": round(base_margin, 2)
                    },
                    "forecast": forecast,
                    "investor_metrics": {
                        "investment_amount": round(investment_amount, 2),
                        "total_profit_5_years": round(total_forecast_profit, 2),
                        "average_annual_profit": round(average_annual_profit, 2),
                        "average_margin": round(average_margin, 2),
                        "roi_5_years": round(roi_5_years, 2),
                        "payback_period": payback_period,
                        "revenue_growth_rate": round(revenue_growth, 2),
                        "expense_growth_rate": round(expense_growth, 2)
                    },
                    "scenario_info": {
                        "name": scenario,
                        "description": ufic_scenario["description"],
                        "detailed_description": ufic_excel_data[scenario].get("detailed_description", {}),
                        "revenue_growth_rate": round(revenue_growth, 2),
                        "expense_growth_rate": round(expense_growth, 2),
                        "cleaners_info": f"Количество уборщиц: {forecast[0]['cleaners_count']} (2026) → {forecast[-1]['cleaners_count']} (2030)"
                    }
                }
            
            # Получаем данные прибылей/убытков за 2025 год
            expense_breakdown_2025 = {}  # Инициализируем для всех случаев
            total_expenses_2025 = 0  # Инициализируем для всех случаев
            
            if company == "ВАШ ДОМ модель":
                result_2025 = await get_consolidated_profit_loss(conn)
                total_expenses_2025 = result_2025["summary"]["total_expenses"]
                
                # Получаем детализацию расходов из consolidated_expenses
                consolidated_expenses = await get_consolidated_expenses(conn)
                
                # Формируем детализацию расходов 2025
                expense_breakdown_2025 = {}
                
                # Проверяем, что consolidated_expenses не None и содержит данные
                if consolidated_expenses and "expenses" in consolidated_expenses:
                    # Проходим по всем категориям расходов
                    for item in consolidated_expenses["expenses"]:
                        category = item["category"]
                        amount = item["amount"]
                        
                        # Нормализуем название категории
                        category_name = category.lower().replace(' ', '_').replace('-', '_')
                        expense_breakdown_2025[category_name] = amount
                
                # Если нет детализации, используем общие расходы
                if not expense_breakdown_2025:
                    expense_breakdown_2025 = {"operating_expenses": total_expenses_2025}
            else:
                # Для ВАШ ДОМ ФАКТ используем специальную логику с тремя сценариями
                # Получаем выручку
                revenue_rows = await conn.fetch("""
                    SELECT month, revenue
                    FROM monthly_revenue
                    WHERE company = $1
                    ORDER BY month
                """, company)
                
                revenue_by_month = {row['month']: float(row['revenue']) for row in revenue_rows}
                
                # Получаем расходы по категориям для детализации (исключаем Ленинск-Кузнецкий)
                expense_by_category = await conn.fetch("""
                    SELECT 
                        category,
                        SUM(amount) as total_amount
                    FROM financial_transactions
                    WHERE type = 'expense' AND company = $1
                      AND category NOT ILIKE '%ленинск%'
                      AND category NOT ILIKE '%кузнец%'
                    GROUP BY category
                    ORDER BY total_amount DESC
                """, company)
                
                # Формируем детализацию расходов по категориям с перераспределением
                expense_breakdown_2025 = {}
                excluded_amount = 0  # Сумма для перераспределения в зарплату
                
                for row in expense_by_category:
                    category = row['category']
                    category_lower = category.lower()
                    amount = float(row['total_amount'])
                    
                    # Исключаемые категории (суммы идут в зарплату)
                    if any(keyword in category_lower for keyword in ['кредит', 'аутсорсинг', 'продукт', 'питание', 'юридическ']):
                        excluded_amount += amount
                        continue
                    
                    # Текущий ремонт: берем только 30%, остальные 70% в зарплату
                    if 'текущий' in category_lower and 'ремонт' in category_lower:
                        excluded_amount += amount * 0.7
                        amount = amount * 0.3
                    
                    category_name = category.lower().replace(' ', '_').replace('-', '_')
                    expense_breakdown_2025[category_name] = amount
                
                # Добавляем исключенные суммы к зарплате
                if 'зарплата' in expense_breakdown_2025:
                    expense_breakdown_2025['зарплата'] += excluded_amount
                else:
                    # Если зарплаты нет как отдельной категории, создаем
                    expense_breakdown_2025['зарплата'] = excluded_amount
                
                # Общие расходы после перераспределения
                total_expenses_2025 = sum(expense_breakdown_2025.values())
                
                # Получаем расходы общие
                expense_rows = await conn.fetch("""
                    SELECT 
                        project as month,
                        SUM(amount) as amount
                    FROM financial_transactions
                    WHERE type = 'expense' AND company = $1
                    GROUP BY project
                """, company)
                
                expenses_by_month = {row['month']: float(row['amount']) for row in expense_rows}
                
                # Формируем данные 2025
                result_2025 = {
                    "profit_loss": [],
                    "summary": {
                        "total_revenue": sum(revenue_by_month.values()),
                        "total_expenses": sum(expenses_by_month.values()),
                        "total_profit": sum(revenue_by_month.values()) - sum(expenses_by_month.values()),
                        "margin": 0
                    }
                }
                
                if result_2025["summary"]["total_revenue"] > 0:
                    result_2025["summary"]["margin"] = round(
                        (result_2025["summary"]["total_profit"] / result_2025["summary"]["total_revenue"] * 100), 2
                    )
            
            # Базовые значения 2025
            base_revenue = result_2025["summary"]["total_revenue"]
            base_expenses = total_expenses_2025  # Используем расходы без Ленинск-Кузнецкий
            base_profit = base_revenue - base_expenses
            
            # Данные "Уборщицы" из УФИЦ модель для интеграции в ВАШ ДОМ модель
            ufic_cleaners_data = {
                "pessimistic": {
                    2026: 34347600,
                    2027: 35996285,
                    2028: 37724106,
                    2029: 39534864,
                    2030: 41432537
                },
                "realistic": {
                    2026: 24615780,
                    2027: 25797337,
                    2028: 27035610,
                    2029: 28333319,
                    2030: 29693318
                },
                "optimistic": {
                    2026: 16028880,
                    2027: 17134873,
                    2028: 18317179,
                    2029: 19581064,
                    2030: 20932158
                }
            }
            
            # Данные "Швеи" и "Аутсорсинг" из УФИЦ модель для вычета из выручки ВАШ ДОМ модель
            # Индексация 4.8% для pessimistic и realistic, 6.9% для optimistic
            ufic_sewing_outsourcing_data = {
                "pessimistic": {
                    # Швеи + Аутсорсинг для пессимистичного (индексация 4.8%)
                    2026: {"sewing": 10696500, "outsourcing": 4410000},
                    2027: {"sewing": 11210092, "outsourcing": 4621680},
                    2028: {"sewing": 11748576, "outsourcing": 4843520},
                    2029: {"sewing": 12312508, "outsourcing": 5076009},
                    2030: {"sewing": 12903108, "outsourcing": 5319657}
                },
                "realistic": {
                    # Швеи + Аутсорсинг для реалистичного (индексация 4.8%)
                    2026: {"sewing": 12033563, "outsourcing": 18522000},
                    2027: {"sewing": 12611174, "outsourcing": 19411056},
                    2028: {"sewing": 13216510, "outsourcing": 20342787},
                    2029: {"sewing": 13850902, "outsourcing": 21319240},
                    2030: {"sewing": 14515745, "outsourcing": 22342563}
                },
                "optimistic": {
                    # Швеи + Аутсорсинг для оптимистичного (индексация 6.9%)
                    2026: {"sewing": 16044750, "outsourcing": 27783000},
                    2027: {"sewing": 17151834, "outsourcing": 29699827},
                    2028: {"sewing": 18335311, "outsourcing": 31748533},
                    2029: {"sewing": 19600447, "outsourcing": 33939194},
                    2030: {"sewing": 20953078, "outsourcing": 36281598}
                }
            }
            
            # Настройки для разных сценариев ВАШ ДОМ ФАКТ (точные данные из Excel)
            vasdom_scenarios = {
                "pessimistic": {
                    "years_data": [
                        {"year": 2026, "revenue": 65539380.00, "expenses": 50923731.00},
                        {"year": 2027, "revenue": 72093318.00, "expenses": 55354095.60},
                        {"year": 2028, "revenue": 79302649.80, "expenses": 60169901.91},
                        {"year": 2029, "revenue": 87232914.78, "expenses": 65404683.38},
                        {"year": 2030, "revenue": 95956206.26, "expenses": 71094890.83},
                    ],
                    "description": "Консервативный прогноз: стабильный уверенный рост, маржа 22-26%",
                    "detailed_description": {
                        "summary": "Консервативный сценарий с минимальным стабильным ростом",
                        "revenue_factors": [
                            "Годовые договора с управляющими компаниями обеспечивают стабильность",
                            "Медленное расширение клиентской базы (+5-7% новых клиентов)",
                            "Улучшение качества сервиса для удержания клиентов"
                        ],
                        "expense_factors": [
                            "Зарплата: основная статья расходов (60-65% от общих расходов)",
                            "Индексация ФОТ на уровне инфляции 7-9%",
                            "Минимальные инвестиции в оборудование"
                        ],
                        "advantages": [
                            "Стабильность обеспечивается за счет годовых договоров с управляющими компаниями",
                            "Минимальные риски за счет консервативного подхода",
                            "Предсказуемый денежный поток и планирование бюджета"
                        ],
                        "disadvantages": [
                            "Зависимость от постоянной клиентской базы",
                            "Ограниченные возможности для масштабирования",
                            "Медленный рост прибыли"
                        ]
                    }
                },
                "realistic": {
                    "years_data": [
                        {"year": 2026, "revenue": 71393318.00, "expenses": 55686788.04},  # маржа ~22%
                        {"year": 2027, "revenue": 80674449.34, "expenses": 60505837.01},  # маржа ~25%
                        {"year": 2028, "revenue": 91162127.75, "expenses": 65636731.98},  # маржа ~28%
                        {"year": 2029, "revenue": 103013204.40, "expenses": 69808683.00},  # маржа ~32%
                        {"year": 2030, "revenue": 116404920.90, "expenses": 75663198.59},  # маржа ~35%
                    ],
                    "description": "Реалистичный прогноз: стабильный уверенный рост, маржа растет с 22% до 35%",
                    "detailed_description": {
                        "summary": "Реалистичный сценарий с уверенным ростом и повышением эффективности",
                        "revenue_factors": [
                            "Расширение портфеля услуг: генеральная уборка, мойка окон, химчистка",
                            "Привлечение корпоративных клиентов (+15-20 B2B контрактов)",
                            "Оптимизация бизнес-процессов и повышение производительности"
                        ],
                        "expense_factors": [
                            "Зарплата: 55-60% от расходов (включая перераспределенные категории)",
                            "Увеличение штата на 10-15% для масштабирования",
                            "Инвестиции в оборудование и технологии: 5-7% от расходов"
                        ],
                        "advantages": [
                            "Сбалансированный подход между ростом и стабильностью",
                            "Постепенное улучшение рентабельности (маржа растет с 22% до 35%)",
                            "Диверсификация клиентской базы за счет B2B контрактов"
                        ],
                        "disadvantages": [
                            "Требуется постоянный контроль качества при расширении",
                            "Необходимость инвестиций в технологии и обучение персонала",
                            "Зависимость от успешного привлечения B2B клиентов"
                        ]
                    }
                },
                "optimistic": {
                    "years_data": [
                        {"year": 2026, "revenue": 75393318.00, "expenses": 58806788.04},  # маржа ~22%
                        {"year": 2027, "revenue": 86702315.70, "expenses": 64381634.69},  # маржа ~26%
                        {"year": 2028, "revenue": 99707663.06, "expenses": 68895414.10},  # маржа ~31%
                        {"year": 2029, "revenue": 114663812.50, "expenses": 72398280.93},  # маржа ~37%
                        {"year": 2030, "revenue": 131863384.40, "expenses": 79118030.64},  # маржа ~40%
                    ],
                    "description": "Оптимистичный прогноз: активный рост, маржа растет с 22% до 40%",
                    "detailed_description": {
                        "summary": "Оптимистичный прогноз с активным расширением и масштабированием",
                        "revenue_factors": [
                            "Активное расширение в новые регионы (3-5 новых городов)",
                            "Запуск premium-услуг: генеральная уборка, дезинфекция, озонирование",
                            "Привлечение крупных федеральных контрактов B2B (+30-40 клиентов)"
                        ],
                        "expense_factors": [
                            "Зарплата: 50-55% от расходов (с учетом автоматизации)",
                            "Масштабирование команды: рост штата на 20-25%",
                            "Интенсивный маркетинг и бренд-реклама: 5-8% от выручки"
                        ],
                        "advantages": [
                            "Максимальный потенциал роста выручки и прибыли",
                            "Быстрое масштабирование и выход на новые рынки",
                            "Создание сильного бренда через premium-услуги"
                        ],
                        "disadvantages": [
                            "Высокие первоначальные инвестиции в IT и маркетинг",
                            "Сложность управления быстрорастущей командой",
                            "Риски при входе на новые региональные рынки"
                        ]
                    }
                }
            }
            
            current_scenario = vasdom_scenarios.get(scenario, vasdom_scenarios["realistic"])
            
            # Генерируем прогноз на 2026-2030 с точными данными из Excel
            forecast = []
            
            # Используем точные данные по годам
            for year_data in current_scenario["years_data"]:
                year = year_data["year"]
                revenue = year_data["revenue"]
                expenses = year_data["expenses"]
                profit = revenue - expenses
                margin = (profit / revenue * 100) if revenue > 0 else 0
                
                # Детализация расходов - пропорционально от базовых расходов 2025
                expense_breakdown_year = {}
                if base_expenses > 0:
                    expenses_multiplier = expenses / base_expenses
                    for category, amount in expense_breakdown_2025.items():
                        expense_breakdown_year[category] = round(amount * expenses_multiplier, 2)
                else:
                    expense_breakdown_year = {"operating_expenses": expenses}
                
                # ДЛЯ "ВАШ ДОМ модель": интегрируем данные "Уборщицы" из УФИЦ модель
                if company == "ВАШ ДОМ модель":
                    cleaners_amount = ufic_cleaners_data.get(scenario, {}).get(year, 0)
                    
                    # Добавляем "Аутсорсинг персонала" с данными уборщиц
                    if "аутсорсинг_персонала" in expense_breakdown_year:
                        expense_breakdown_year["аутсорсинг_персонала"] += cleaners_amount
                    else:
                        expense_breakdown_year["аутсорсинг_персонала"] = cleaners_amount
                    
                    # Уменьшаем ФОТ или Зарплату на сумму уборщиц
                    # Ищем категорию ФОТ или Зарплата
                    fot_key = None
                    for key in expense_breakdown_year.keys():
                        if "фот" in key.lower() or "зарплата" in key.lower():
                            fot_key = key
                            break
                    
                    if fot_key and expense_breakdown_year[fot_key] > cleaners_amount:
                        expense_breakdown_year[fot_key] -= cleaners_amount
                    elif fot_key:
                        # Если ФОТ меньше чем сумма уборщиц, обнуляем
                        expense_breakdown_year[fot_key] = 0
                
                # Детализация доходов
                revenue_breakdown_year = {
                    "main_revenue": round(revenue, 2)
                }
                
                # ДЛЯ "ВАШ ДОМ модель": вычитаем "Швеи" и "Аутсорсинг" из УФИЦ модель
                if company == "ВАШ ДОМ модель":
                    ufic_data_year = ufic_sewing_outsourcing_data.get(scenario, {}).get(year, {})
                    sewing_amount = ufic_data_year.get("sewing", 0)
                    outsourcing_amount = ufic_data_year.get("outsourcing", 0)
                    
                    # Вычитаем из общей выручки
                    net_revenue = revenue - sewing_amount - outsourcing_amount
                    
                    revenue_breakdown_year = {
                        "vasdom_revenue": round(net_revenue, 2),
                        "ufic_sewing": round(sewing_amount, 2),
                        "ufic_outsourcing": round(outsourcing_amount, 2),
                        "total": round(revenue, 2)
                    }
                
                forecast.append({
                    "year": year,
                    "revenue": round(revenue, 2),
                    "expenses": round(expenses, 2),
                    "profit": round(profit, 2),
                    "margin": round(margin, 2),
                    "revenue_breakdown": revenue_breakdown_year,
                    "expense_breakdown": expense_breakdown_year
                })
            
            # Расчеты для инвестора
            total_forecast_profit = sum(f["profit"] for f in forecast)
            average_annual_profit = total_forecast_profit / len(forecast)
            
            # Средняя рентабельность
            average_margin = sum(f["margin"] for f in forecast) / len(forecast)
            
            # ROI за 5 лет (инвестиции = 40 млн для ВАШ ДОМ+УФИЦ)
            investment_amount = 40000000  # 40 млн рублей
            roi_5_years = (total_forecast_profit / investment_amount * 100) if investment_amount > 0 else 0
            
            # Срок окупаемости (в годах)
            cumulative_profit = 0
            payback_period = None
            for i, f in enumerate(forecast):
                cumulative_profit += f["profit"]
                if cumulative_profit >= investment_amount and payback_period is None:
                    payback_period = i + 1
            
            if payback_period is None:
                payback_period = "> 5 лет"
            
            investor_metrics = {
                "investment_amount": round(investment_amount, 2),
                "total_profit_5_years": round(total_forecast_profit, 2),
                "average_annual_profit": round(average_annual_profit, 2),
                "average_margin": round(average_margin, 2),
                "roi_5_years": round(roi_5_years, 2),
                "payback_period": payback_period,
                "revenue_growth_rate": round(((forecast[0]["revenue"] / base_revenue) - 1) * 100, 2),
                "expense_growth_rate": round(((forecast[0]["expenses"] / base_expenses) - 1) * 100, 2)
            }
            
            return {
                "company": company,
                "scenario": scenario,
                "base_year": 2025,
                "base_data": {
                    "revenue": round(base_revenue, 2),
                    "expenses": round(base_expenses, 2),
                    "profit": round(base_profit, 2),
                    "margin": result_2025["summary"]["margin"]
                },
                "forecast": forecast,
                "investor_metrics": investor_metrics,
                "scenario_info": {
                    "name": scenario,
                    "description": current_scenario.get("description", scenario_config.get("description", "")),
                    "detailed_description": current_scenario.get("detailed_description", {}),
                    "revenue_growth_rate": round(((forecast[0]["revenue"] / base_revenue) - 1) * 100, 2),
                    "expense_growth_rate": round(((forecast[0]["expenses"] / base_expenses) - 1) * 100, 2)
                }
            }
            
        finally:
            await conn.close()
    except Exception as e:
        logger.error(f"Error calculating forecast: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/finances/export-all")
async def export_all_financial_data():
    """
    Экспорт финансовых данных анализа в XLSX файл:
    - Анализ - Выручка (по всем компаниям)
    - Анализ - Расходы (по всем компаниям)
    
    Примечание: Экспорт прогнозов доступен отдельно на странице "Прогноз 26-30"
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        wb.remove(wb.active)
        
        conn = await get_db_connection()
        
        try:
            companies = ["ВАШ ДОМ ФАКТ", "УФИЦ модель", "ВАШ ДОМ модель"]
            
            # === ЛИСТ 1: ВЫРУЧКА ===
            ws_revenue = wb.create_sheet("Анализ - Выручка")
            ws_revenue.append(["Анализ выручки по компаниям"])
            ws_revenue.merge_cells('A1:D1')
            ws_revenue['A1'].font = Font(bold=True, size=14)
            ws_revenue['A1'].alignment = Alignment(horizontal="center")
            ws_revenue.append([])
            
            header = ["Компания", "Месяц", "Сумма (₽)", ""]
            ws_revenue.append(header)
            
            for i, cell in enumerate(ws_revenue[3], 1):
                if i <= 3:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            for company in companies:
                query = """
                    SELECT 
                        TO_CHAR(date, 'TMMonth YYYY') as month,
                        SUM(amount) as total
                    FROM financial_transactions
                    WHERE type = 'income' AND company = $1
                    GROUP BY TO_CHAR(date, 'TMMonth YYYY'), DATE_TRUNC('month', date)
                    ORDER BY DATE_TRUNC('month', date)
                """
                rows = await conn.fetch(query, company)
                
                total_company = 0
                for row in rows:
                    month = row['month'].strip()
                    amount = float(row['total'])
                    total_company += amount
                    ws_revenue.append([company, month, amount, ""])
                
                ws_revenue.append([company, "ИТОГО", total_company, ""])
                last_row = ws_revenue.max_row
                ws_revenue.cell(last_row, 1).font = Font(bold=True)
                ws_revenue.cell(last_row, 2).font = Font(bold=True)
                ws_revenue.cell(last_row, 3).font = Font(bold=True)
                ws_revenue.cell(last_row, 3).fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                ws_revenue.append([])
            
            for col in range(1, 4):
                ws_revenue.column_dimensions[get_column_letter(col)].width = 25
            
            # === ЛИСТ 2: РАСХОДЫ ===
            ws_expense = wb.create_sheet("Анализ - Расходы")
            ws_expense.append(["Анализ расходов по компаниям"])
            ws_expense.merge_cells('A1:E1')
            ws_expense['A1'].font = Font(bold=True, size=14)
            ws_expense['A1'].alignment = Alignment(horizontal="center")
            ws_expense.append([])
            
            header = ["Компания", "Категория", "Сумма (₽)", "Процент (%)", ""]
            ws_expense.append(header)
            
            for i, cell in enumerate(ws_expense[3], 1):
                if i <= 4:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            for company in companies:
                query = """
                    SELECT 
                        category,
                        SUM(amount) as total
                    FROM financial_transactions
                    WHERE type = 'expense' AND company = $1
                    GROUP BY category
                    ORDER BY total DESC
                """
                rows = await conn.fetch(query, company)
                
                total_company = sum(float(row['total']) for row in rows)
                
                for row in rows:
                    category = row['category']
                    amount = float(row['total'])
                    percentage = round((amount / total_company * 100), 2) if total_company > 0 else 0
                    ws_expense.append([company, category, amount, percentage, ""])
                
                ws_expense.append([company, "ИТОГО", total_company, 100.0, ""])
                last_row = ws_expense.max_row
                for col in range(1, 5):
                    ws_expense.cell(last_row, col).font = Font(bold=True)
                    ws_expense.cell(last_row, col).fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                ws_expense.append([])
            
            for col in range(1, 5):
                ws_expense.column_dimensions[get_column_letter(col)].width = 25
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"financial_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        finally:
            if conn:
                await conn.close()
        
    except Exception as e:
        logger.error(f"Error exporting all financial data: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


@router.get("/finances/export-forecast")
async def export_forecast_data(
    company: Optional[str] = "ВАШ ДОМ ФАКТ",
    scenario: Optional[str] = "realistic"
):
    """
    Экспорт прогноза в XLSX для конкретной модели и сценария
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        # Получаем данные прогноза
        forecast_data = await get_forecast(company=company, scenario=scenario)
        
        if not forecast_data or 'forecast' not in forecast_data:
            raise HTTPException(status_code=404, detail="Forecast data not found")
        
        # Создаем workbook
        wb = Workbook()
        ws = wb.active
        
        # Определяем название компании для отображения
        company_display = {
            "ВАШ ДОМ ФАКТ": "ВАШ ДОМ+УФИЦ",
            "УФИЦ модель": "УФИЦ модель",
            "ВАШ ДОМ модель": "ВАШ ДОМ модель"
        }.get(company, company)
        
        scenario_name = {
            "pessimistic": "Пессимистичный",
            "realistic": "Реалистичный",
            "optimistic": "Оптимистичный"
        }.get(scenario, scenario)
        
        ws.title = f"{company_display} - {scenario_name[:20]}"
        
        # Заголовок
        ws.append([f"Прогноз: {company_display} - {scenario_name}"])
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.append([])
        
        # Базовый год 2025
        ws.append(["БАЗОВЫЙ ГОД 2025"])
        ws.merge_cells(f'A{ws.max_row}:B{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        ws[f'A{ws.max_row}'].fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
        
        ws.append(["Выручка", forecast_data['base_data']['revenue'], "", "", ""])
        ws.append(["Расходы", forecast_data['base_data']['expenses'], "", "", ""])
        ws.append(["Прибыль", forecast_data['base_data']['profit'], "", "", ""])
        ws.append(["Маржа (%)", forecast_data['base_data']['margin'], "", "", ""])
        ws.append([])
        
        # Прогноз 2026-2030
        ws.append(["ПРОГНОЗ 2026-2030"])
        ws.merge_cells(f'A{ws.max_row}:E{ws.max_row}')
        ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
        ws[f'A{ws.max_row}'].fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
        
        header = ["Год", "Выручка (₽)", "Расходы (₽)", "Прибыль (₽)", "Маржа (%)"]
        ws.append(header)
        
        header_row = ws.max_row
        for col_idx in range(1, 6):
            cell = ws.cell(header_row, col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        total_revenue = 0
        total_expenses = 0
        total_profit = 0
        
        for year_data in forecast_data['forecast']:
            ws.append([
                year_data['year'],
                year_data['revenue'],
                year_data['expenses'],
                year_data['profit'],
                year_data['margin']
            ])
            total_revenue += year_data['revenue']
            total_expenses += year_data['expenses']
            total_profit += year_data['profit']
        
        # Итого за 5 лет
        ws.append([])
        avg_margin = round(total_profit / total_revenue * 100, 2) if total_revenue > 0 else 0
        ws.append(["ИТОГО 5 лет", total_revenue, total_expenses, total_profit, avg_margin])
        last_row = ws.max_row
        for col in range(1, 6):
            ws.cell(last_row, col).font = Font(bold=True)
            ws.cell(last_row, col).fill = PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid")
        
        # Метрики для инвестора
        if 'investor_metrics' in forecast_data:
            ws.append([])
            ws.append(["РАСЧЕТЫ ДЛЯ ИНВЕСТОРА"])
            ws.merge_cells(f'A{ws.max_row}:E{ws.max_row}')
            ws[f'A{ws.max_row}'].font = Font(bold=True, size=12)
            ws[f'A{ws.max_row}'].fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
            
            metrics = forecast_data['investor_metrics']
            ws.append(["Инвестиции", metrics.get('investment_amount', 0), "", "", ""])
            ws.append(["Прибыль за 5 лет", metrics.get('total_profit_5_years', 0), "", "", ""])
            ws.append(["Средняя прибыль/год", metrics.get('average_annual_profit', 0), "", "", ""])
            ws.append(["Средняя маржа (%)", round(metrics.get('average_margin', 0), 2), "", "", ""])
            ws.append(["ROI за 5 лет (%)", round(metrics.get('roi_5_years', 0), 2), "", "", ""])
            ws.append(["Срок окупаемости", str(metrics.get('payback_period', 'N/A')), "", "", ""])
        
        # Автоширина колонок
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Сохраняем в BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Формируем имя файла
        company_short = company_display.replace(" ", "_").replace("+", "")
        scenario_short = scenario_name
        filename = f"forecast_{company_short}_{scenario_short}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        logger.error(f"Error exporting forecast: {e}")
        raise HTTPException(status_code=500, detail=f"Export failed: {str(e)}")


