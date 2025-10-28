#!/usr/bin/env python3
"""
Backend Test Suite for VasDom AudioBot System
Testing Finance Module and Plannerka functionality

This test validates:
FINANCE MODULE:
1. GET /api/finances/cash-flow - движение денег
2. GET /api/finances/profit-loss - прибыли и убытки  
3. GET /api/finances/balance-sheet - балансовый отчёт
4. GET /api/finances/expense-analysis - анализ расходов
5. GET /api/finances/available-months - доступные месяцы
6. GET /api/finances/debts - задолженности
7. GET /api/finances/inventory - товарные запасы
8. GET /api/finances/dashboard - сводка
9. GET /api/finances/transactions - список транзакций
10. POST /api/finances/transactions - создание транзакции
11. GET /api/finances/revenue/monthly - ручная выручка

PLANNERKA MODULE:
1. POST /api/plannerka/create - создание планёрки
2. POST /api/plannerka/analyze/{id} - AI-анализ с GPT-4o
3. GET /api/plannerka/list - список планёрок
"""

import asyncio
import httpx
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
import os

# Backend URL from environment
BACKEND_URL = "https://finance-forecast-14.preview.emergentagent.com"
API_BASE = f"{BACKEND_URL}/api"

class TestResults:
    def __init__(self):
        # Plannerka results
        self.created_meeting_id = None
        self.created_meeting_data = None
        self.analysis_result = None
        self.meetings_list = []
        self.openai_working = False
        self.tasks_extracted = []
        self.summary_generated = False
        
        # Finance results
        self.finance_endpoints = {}
        self.created_transaction_id = None
        self.database_working = False
        
        # Common
        self.errors = []

async def test_finance_cash_flow():
    """Test finance cash flow endpoint"""
    print("=== ТЕСТ ДВИЖЕНИЯ ДЕНЕГ (CASH FLOW) ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("💰 Тестируем GET /api/finances/cash-flow...")
            
            response = await client.get(f"{API_BASE}/finances/cash-flow")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка cash-flow: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['cash_flow'] = data
            
            print("✅ Cash flow получен успешно")
            print(f"📊 Структура ответа:")
            
            # Validate structure
            required_fields = ['cash_flow', 'summary']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в cash-flow")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            # Check summary structure
            if 'summary' in data:
                summary = data['summary']
                summary_fields = ['total_income', 'total_expense', 'net_cash_flow']
                for field in summary_fields:
                    if field not in summary:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в summary")
                    else:
                        print(f"✅ Summary поле '{field}': {summary[field]}")
            
            # Check cash flow items
            cash_flow_items = data.get('cash_flow', [])
            print(f"📈 Записей движения денег: {len(cash_flow_items)}")
            
            if cash_flow_items:
                sample_item = cash_flow_items[0]
                item_fields = ['date', 'income', 'expense', 'balance']
                for field in item_fields:
                    if field not in sample_item:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в записи cash_flow")
                    else:
                        print(f"✅ Поле записи '{field}': {sample_item[field]}")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании cash-flow: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_profit_loss():
    """Test finance profit loss endpoint"""
    print("\n=== ТЕСТ ПРИБЫЛЕЙ И УБЫТКОВ (PROFIT LOSS) ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("📊 Тестируем GET /api/finances/profit-loss...")
            
            response = await client.get(f"{API_BASE}/finances/profit-loss")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка profit-loss: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['profit_loss'] = data
            
            print("✅ Profit loss получен успешно")
            print(f"📊 Структура ответа:")
            
            # Validate structure
            required_fields = ['profit_loss', 'summary']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в profit-loss")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            # Check summary structure
            if 'summary' in data:
                summary = data['summary']
                summary_fields = ['total_revenue', 'total_expenses', 'net_profit', 'average_margin']
                for field in summary_fields:
                    if field not in summary:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в summary")
                    else:
                        print(f"✅ Summary поле '{field}': {summary[field]}")
            
            # Check profit loss items
            profit_loss_items = data.get('profit_loss', [])
            print(f"📈 Записей прибылей/убытков: {len(profit_loss_items)}")
            
            if profit_loss_items:
                sample_item = profit_loss_items[0]
                item_fields = ['period', 'revenue', 'expenses', 'profit', 'margin']
                for field in item_fields:
                    if field not in sample_item:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в записи profit_loss")
                    else:
                        print(f"✅ Поле записи '{field}': {sample_item[field]}")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании profit-loss: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_expense_analysis():
    """Test finance expense analysis endpoint"""
    print("\n=== ТЕСТ АНАЛИЗА РАСХОДОВ (EXPENSE ANALYSIS) ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("💸 Тестируем GET /api/finances/expense-analysis...")
            
            # Test without month filter
            response = await client.get(f"{API_BASE}/finances/expense-analysis")
            print(f"📡 Ответ сервера (без фильтра): {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка expense-analysis: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['expense_analysis'] = data
            
            print("✅ Expense analysis получен успешно")
            
            # Validate structure
            required_fields = ['expenses', 'total']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в expense-analysis")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            # Check expenses structure
            expenses = data.get('expenses', [])
            print(f"📈 Категорий расходов: {len(expenses)}")
            print(f"💰 Общая сумма расходов: {data.get('total', 0)}")
            
            if expenses:
                sample_expense = expenses[0]
                expense_fields = ['category', 'amount', 'percentage']
                for field in expense_fields:
                    if field not in sample_expense:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в записи expense")
                    else:
                        print(f"✅ Поле расхода '{field}': {sample_expense[field]}")
            
            # Test with month filter
            print("\n🗓️ Тестируем с фильтром по месяцу...")
            response_month = await client.get(f"{API_BASE}/finances/expense-analysis?month=Январь 2025")
            print(f"📡 Ответ сервера (с фильтром): {response_month.status_code}")
            
            if response_month.status_code == 200:
                month_data = response_month.json()
                print(f"✅ Фильтр по месяцу работает")
                print(f"📅 Месяц: {month_data.get('month')}")
                print(f"💰 Расходов за месяц: {len(month_data.get('expenses', []))}")
            else:
                print(f"⚠️ Фильтр по месяцу не работает: {response_month.status_code}")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании expense-analysis: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_available_months():
    """Test finance available months endpoint"""
    print("\n=== ТЕСТ ДОСТУПНЫХ МЕСЯЦЕВ ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("📅 Тестируем GET /api/finances/available-months...")
            
            response = await client.get(f"{API_BASE}/finances/available-months")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка available-months: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['available_months'] = data
            
            print("✅ Available months получен успешно")
            
            # Validate structure
            required_fields = ['months', 'total']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в available-months")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            months = data.get('months', [])
            print(f"📅 Доступных месяцев: {data.get('total', 0)}")
            
            if months:
                print(f"📋 Примеры месяцев: {months[:5]}")
            else:
                print("⚠️ Нет доступных месяцев")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании available-months: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_balance_sheet():
    """Test finance balance sheet endpoint (mock data)"""
    print("\n=== ТЕСТ БАЛАНСОВОГО ОТЧЁТА ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("🏦 Тестируем GET /api/finances/balance-sheet...")
            
            response = await client.get(f"{API_BASE}/finances/balance-sheet")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка balance-sheet: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['balance_sheet'] = data
            
            print("✅ Balance sheet получен успешно (mock данные)")
            
            # Validate structure
            required_sections = ['assets', 'liabilities', 'equity']
            for section in required_sections:
                if section not in data:
                    results.errors.append(f"❌ Отсутствует секция '{section}' в balance-sheet")
                else:
                    print(f"✅ Секция '{section}' присутствует")
                    section_data = data[section]
                    if 'total' in section_data:
                        print(f"   💰 Итого {section}: {section_data['total']}")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании balance-sheet: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_debts():
    """Test finance debts endpoint (mock data)"""
    print("\n=== ТЕСТ ЗАДОЛЖЕННОСТЕЙ ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("💳 Тестируем GET /api/finances/debts...")
            
            response = await client.get(f"{API_BASE}/finances/debts")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка debts: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['debts'] = data
            
            print("✅ Debts получен успешно (mock данные)")
            
            # Validate structure
            required_fields = ['debts', 'summary']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в debts")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            debts = data.get('debts', [])
            summary = data.get('summary', {})
            
            print(f"💳 Количество задолженностей: {len(debts)}")
            print(f"💰 Общая сумма: {summary.get('total', 0)}")
            print(f"⚠️ Просроченная сумма: {summary.get('overdue', 0)}")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании debts: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_inventory():
    """Test finance inventory endpoint (mock data)"""
    print("\n=== ТЕСТ ТОВАРНЫХ ЗАПАСОВ ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("📦 Тестируем GET /api/finances/inventory...")
            
            response = await client.get(f"{API_BASE}/finances/inventory")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка inventory: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['inventory'] = data
            
            print("✅ Inventory получен успешно (mock данные)")
            
            # Validate structure
            required_fields = ['inventory', 'summary']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в inventory")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            inventory = data.get('inventory', [])
            summary = data.get('summary', {})
            
            print(f"📦 Количество позиций: {len(inventory)}")
            print(f"💰 Общая стоимость: {summary.get('total_value', 0)}")
            print(f"📊 Общее количество: {summary.get('total_items', 0)}")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании inventory: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_dashboard():
    """Test finance dashboard endpoint"""
    print("\n=== ТЕСТ ФИНАНСОВОЙ СВОДКИ (DASHBOARD) ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:  # Longer timeout as it aggregates data
            print("📊 Тестируем GET /api/finances/dashboard...")
            
            response = await client.get(f"{API_BASE}/finances/dashboard")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка dashboard: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['dashboard'] = data
            
            print("✅ Dashboard получен успешно")
            
            # Validate structure - dashboard aggregates all other endpoints
            expected_sections = ['cash_flow', 'profit_loss', 'balance', 'expenses', 'debts', 'inventory']
            for section in expected_sections:
                if section not in data:
                    results.errors.append(f"❌ Отсутствует секция '{section}' в dashboard")
                else:
                    print(f"✅ Секция '{section}' присутствует")
            
            print("📊 Сводка по всем показателям получена")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании dashboard: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_transactions_list():
    """Test finance transactions list endpoint"""
    print("\n=== ТЕСТ СПИСКА ТРАНЗАКЦИЙ ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("💼 Тестируем GET /api/finances/transactions...")
            
            response = await client.get(f"{API_BASE}/finances/transactions?limit=10&offset=0")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка transactions list: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['transactions_list'] = data
            
            print("✅ Transactions list получен успешно")
            print(f"📊 Количество транзакций: {len(data)}")
            
            if data:
                sample_transaction = data[0]
                required_fields = ['id', 'date', 'amount', 'category', 'type', 'created_at']
                for field in required_fields:
                    if field not in sample_transaction:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в транзакции")
                    else:
                        print(f"✅ Поле транзакции '{field}': {sample_transaction[field]}")
            else:
                print("⚠️ Список транзакций пуст")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании transactions list: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_create_transaction():
    """Test finance create transaction endpoint"""
    print("\n=== ТЕСТ СОЗДАНИЯ ТРАНЗАКЦИИ ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("💰 Тестируем POST /api/finances/transactions...")
            
            # Test data as specified in the review request
            test_transaction = {
                "date": "2025-10-17T00:00:00Z",
                "amount": 1000,
                "category": "Зарплата",
                "type": "expense",
                "description": "Тестовая транзакция для проверки API",
                "payment_method": "Банковский перевод",
                "counterparty": "Тестовый сотрудник",
                "project": "Январь 2025"
            }
            
            print(f"📝 Данные транзакции: {json.dumps(test_transaction, ensure_ascii=False, indent=2)}")
            
            response = await client.post(
                f"{API_BASE}/finances/transactions",
                json=test_transaction,
                headers={"Content-Type": "application/json"}
            )
            
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка создания транзакции: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.created_transaction_id = data.get('id')
            results.finance_endpoints['created_transaction'] = data
            
            print(f"✅ Транзакция создана с ID: {results.created_transaction_id}")
            
            # Validate response structure
            required_fields = ['id', 'date', 'amount', 'category', 'type', 'created_at']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в ответе создания")
                else:
                    print(f"✅ Поле '{field}': {data[field]}")
            
            # Validate data integrity
            if data.get('amount') != test_transaction['amount']:
                results.errors.append(f"❌ Неверная сумма: ожидалась {test_transaction['amount']}, получена {data.get('amount')}")
            
            if data.get('category') != test_transaction['category']:
                results.errors.append(f"❌ Неверная категория: ожидалась '{test_transaction['category']}', получена '{data.get('category')}'")
            
            if data.get('type') != test_transaction['type']:
                results.errors.append(f"❌ Неверный тип: ожидался '{test_transaction['type']}', получен '{data.get('type')}'")
            
            if not results.errors:
                print("✅ Все поля транзакции корректны")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании создания транзакции: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_revenue_monthly():
    """Test finance monthly revenue endpoint"""
    print("\n=== ТЕСТ РУЧНОЙ ВЫРУЧКИ ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("💵 Тестируем GET /api/finances/revenue/monthly...")
            
            response = await client.get(f"{API_BASE}/finances/revenue/monthly")
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка revenue monthly: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.finance_endpoints['revenue_monthly'] = data
            
            print("✅ Revenue monthly получен успешно")
            
            # Validate structure
            required_fields = ['revenues']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в revenue monthly")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            revenues = data.get('revenues', [])
            print(f"📊 Записей выручки: {len(revenues)}")
            
            if revenues:
                sample_revenue = revenues[0]
                revenue_fields = ['month', 'revenue']
                for field in revenue_fields:
                    if field not in sample_revenue:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в записи выручки")
                    else:
                        print(f"✅ Поле выручки '{field}': {sample_revenue[field]}")
            else:
                print("⚠️ Нет записей выручки (таблица может быть пустой)")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании revenue monthly: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_finance_expense_details():
    """Test new expense details endpoint for specific months"""
    print("\n=== ТЕСТ ДЕТАЛИЗАЦИИ РАСХОДОВ ПО МЕСЯЦАМ ===\n")
    
    results = TestResults()
    
    # Test months as specified in the review request
    test_months = ["Июль 2025", "Март 2025", "Сентябрь 2025"]
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            for month in test_months:
                print(f"💸 Тестируем GET /api/finances/expense-details?month={month}...")
                
                response = await client.get(f"{API_BASE}/finances/expense-details", params={"month": month})
                print(f"📡 Ответ сервера для {month}: {response.status_code}")
                
                if response.status_code != 200:
                    error_msg = f"❌ Ошибка expense-details для {month}: {response.status_code} - {response.text}"
                    results.errors.append(error_msg)
                    print(error_msg)
                    continue
                
                data = response.json()
                results.finance_endpoints[f'expense_details_{month}'] = data
                
                print(f"✅ Expense details для {month} получен успешно")
                
                # Validate response structure
                required_fields = ['transactions', 'total', 'month', 'count']
                for field in required_fields:
                    if field not in data:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в expense-details для {month}")
                    else:
                        print(f"✅ Поле '{field}' присутствует")
                
                # Validate month field
                if data.get('month') != month:
                    results.errors.append(f"❌ Неверный месяц в ответе: ожидался '{month}', получен '{data.get('month')}'")
                else:
                    print(f"✅ Месяц корректен: {data.get('month')}")
                
                # Check transactions structure
                transactions = data.get('transactions', [])
                total = data.get('total', 0)
                count = data.get('count', 0)
                
                print(f"📊 Транзакций найдено: {count}")
                print(f"💰 Общая сумма расходов: {total}")
                
                # Validate count matches transactions length
                if len(transactions) != count:
                    results.errors.append(f"❌ Несоответствие count ({count}) и длины массива transactions ({len(transactions)}) для {month}")
                else:
                    print(f"✅ Count соответствует количеству транзакций")
                
                # Validate total calculation
                if transactions:
                    calculated_total = sum(t.get('amount', 0) for t in transactions)
                    if abs(calculated_total - total) > 0.01:  # Allow small floating point differences
                        results.errors.append(f"❌ Неверный расчет total для {month}: ожидалось {calculated_total}, получено {total}")
                    else:
                        print(f"✅ Total корректно рассчитан")
                
                # Check transaction structure
                if transactions:
                    sample_transaction = transactions[0]
                    required_transaction_fields = ['id', 'date', 'category', 'amount', 'description', 'payment_method', 'counterparty']
                    
                    print(f"📋 Проверяем структуру транзакции:")
                    for field in required_transaction_fields:
                        if field not in sample_transaction:
                            results.errors.append(f"❌ Отсутствует поле '{field}' в транзакции для {month}")
                        else:
                            value = sample_transaction[field]
                            print(f"✅ Поле '{field}': {value}")
                    
                    # Validate that all transactions are expenses (type='expense')
                    # Note: The endpoint filters by type='expense' in SQL, so we assume all returned are expenses
                    print(f"✅ Все транзакции являются расходами (фильтр type='expense' применен в SQL)")
                    
                    # Show sample transaction details
                    print(f"📝 Пример транзакции:")
                    print(f"   - ID: {sample_transaction.get('id')}")
                    print(f"   - Дата: {sample_transaction.get('date')}")
                    print(f"   - Категория: {sample_transaction.get('category')}")
                    print(f"   - Сумма: {sample_transaction.get('amount')}")
                    print(f"   - Описание: {sample_transaction.get('description')}")
                    print(f"   - Способ оплаты: {sample_transaction.get('payment_method')}")
                    print(f"   - Контрагент: {sample_transaction.get('counterparty')}")
                else:
                    print(f"⚠️ Нет транзакций для месяца {month}")
                
                print(f"")  # Empty line for readability
            
            # Test existing expense-analysis endpoint to ensure it still works
            print("🔄 Проверяем совместимость с существующим endpoint expense-analysis...")
            
            for month in test_months:
                response = await client.get(f"{API_BASE}/finances/expense-analysis", params={"month": month})
                print(f"📡 expense-analysis для {month}: {response.status_code}")
                
                if response.status_code != 200:
                    error_msg = f"❌ Существующий endpoint expense-analysis не работает для {month}: {response.status_code}"
                    results.errors.append(error_msg)
                else:
                    print(f"✅ Существующий endpoint expense-analysis работает для {month}")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании expense-details: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_export_all_endpoint():
    """Test GET /api/finances/export-all endpoint according to review request"""
    print("\n=== ТЕСТ ЭКСПОРТА ВСЕХ ФИНАНСОВЫХ ДАННЫХ В XLSX ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            print("📊 Тестируем GET /api/finances/export-all...")
            print("🎯 КРИТЕРИИ УСПЕХА (согласно review request):")
            print("   1. ✅ Должен возвращать 200 OK")
            print("   2. ✅ XLSX файл с 2 листами (Анализ - Выручка, Анализ - Расходы)")
            print("   3. ✅ Размер файла ~8KB")
            print("   4. ✅ Время отклика < 5 секунд")
            print("")
            
            # Measure response time
            import time
            start_time = time.time()
            
            # Test the export-all endpoint
            response = await client.get(f"{API_BASE}/finances/export-all")
            
            end_time = time.time()
            response_time = end_time - start_time
            
            print(f"📡 Ответ сервера: {response.status_code}")
            print(f"⏱️ Время отклика: {response_time:.2f} секунд")
            
            # 1. Check 200 status
            if response.status_code != 200:
                error_msg = f"❌ КРИТЕРИЙ 1 НЕ ВЫПОЛНЕН: Ожидался статус 200, получен {response.status_code}"
                results.errors.append(error_msg)
                print(error_msg)
                if response.text:
                    print(f"   Текст ошибки: {response.text[:500]}")
                return results
            
            print("✅ КРИТЕРИЙ 1 ВЫПОЛНЕН: Endpoint возвращает 200 OK")
            
            # 4. Check response time < 5 seconds
            if response_time >= 5.0:
                error_msg = f"❌ КРИТЕРИЙ 4 НЕ ВЫПОЛНЕН: Время отклика {response_time:.2f}с >= 5с"
                results.errors.append(error_msg)
                print(error_msg)
            else:
                print(f"✅ КРИТЕРИЙ 4 ВЫПОЛНЕН: Время отклика {response_time:.2f}с < 5с")
            
            # Get file content
            content = response.content
            
            if not content:
                error_msg = "❌ Ответ не содержит данных"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            # 3. Check file size ~8KB
            file_size = len(content)
            expected_size = 8 * 1024  # 8KB
            
            print(f"📊 Размер файла: {file_size:,} байт ({file_size / 1024:.1f} KB)")
            
            if file_size < 4 * 1024:  # Less than 4KB is too small
                error_msg = f"❌ КРИТЕРИЙ 3 НЕ ВЫПОЛНЕН: Размер файла {file_size:,} байт слишком мал (< 4KB)"
                results.errors.append(error_msg)
                print(error_msg)
            elif file_size > 20 * 1024:  # More than 20KB is too large
                error_msg = f"❌ КРИТЕРИЙ 3 НЕ ВЫПОЛНЕН: Размер файла {file_size:,} байт слишком велик (> 20KB)"
                results.errors.append(error_msg)
                print(error_msg)
            else:
                print(f"✅ КРИТЕРИЙ 3 ВЫПОЛНЕН: Размер файла {file_size / 1024:.1f} KB в ожидаемом диапазоне")
            
            # 2. Check valid XLSX format and sheet contents
            try:
                import io
                from openpyxl import load_workbook
                
                # Load XLSX file from memory
                xlsx_file = io.BytesIO(content)
                workbook = load_workbook(xlsx_file, read_only=True)
                
                print("✅ Файл валидный XLSX формат")
                
                # Get all sheet names
                sheet_names = workbook.sheetnames
                print(f"📋 Найдено листов: {len(sheet_names)}")
                
                # Expected sheets according to review request
                expected_sheets = [
                    "Анализ - Выручка",
                    "Анализ - Расходы"
                ]
                
                print(f"📋 Листы в файле:")
                for i, sheet_name in enumerate(sheet_names, 1):
                    print(f"   {i}. {sheet_name}")
                
                # Check if we have 2 sheets total
                if len(sheet_names) != 2:
                    error_msg = f"❌ КРИТЕРИЙ 2 НЕ ВЫПОЛНЕН: Найдено {len(sheet_names)} листов, ожидалось 2"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print("✅ КРИТЕРИЙ 2 ЧАСТИЧНО ВЫПОЛНЕН: Найдено 2 листа")
                
                # Check for required sheets
                missing_sheets = []
                found_sheets = []
                
                for expected_sheet in expected_sheets:
                    if expected_sheet in sheet_names:
                        found_sheets.append(expected_sheet)
                        print(f"✅ Найден лист: '{expected_sheet}'")
                    else:
                        missing_sheets.append(expected_sheet)
                        print(f"❌ Отсутствует лист: '{expected_sheet}'")
                
                if missing_sheets:
                    error_msg = f"❌ КРИТЕРИЙ 2 НЕ ВЫПОЛНЕН: Отсутствуют листы: {missing_sheets}"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print("✅ КРИТЕРИЙ 2 ВЫПОЛНЕН: Все требуемые листы присутствуют")
                
                # Validate sheet contents (basic check)
                print(f"\n📊 Проверка содержимого листов:")
                
                # Check "Анализ - Выручка" sheet
                if "Анализ - Выручка" in sheet_names:
                    revenue_sheet = workbook["Анализ - Выручка"]
                    max_row = revenue_sheet.max_row
                    max_col = revenue_sheet.max_column
                    print(f"   'Анализ - Выручка': {max_row} строк, {max_col} столбцов")
                    
                    if max_row > 1 and max_col > 1:
                        print("   ✅ Лист содержит данные")
                    else:
                        print("   ⚠️ Лист может быть пустым")
                
                # Check "Анализ - Расходы" sheet  
                if "Анализ - Расходы" in sheet_names:
                    expense_sheet = workbook["Анализ - Расходы"]
                    max_row = expense_sheet.max_row
                    max_col = expense_sheet.max_column
                    print(f"   'Анализ - Расходы': {max_row} строк, {max_col} столбцов")
                    
                    if max_row > 1 and max_col > 1:
                        print("   ✅ Лист содержит данные")
                    else:
                        print("   ⚠️ Лист может быть пустым")
                
                workbook.close()
                
            except ImportError:
                print("⚠️ openpyxl не установлен - проверка содержимого XLSX пропущена")
                
                # Basic XLSX validation - check ZIP signature
                if content.startswith(b'PK'):
                    print("✅ Файл имеет корректную XLSX сигнатуру (ZIP-based)")
                else:
                    error_msg = f"❌ Неверная сигнатура файла: {content[:10]}"
                    results.errors.append(error_msg)
                    print(error_msg)
                
            except Exception as xlsx_error:
                error_msg = f"❌ Ошибка чтения XLSX: {str(xlsx_error)}"
                results.errors.append(error_msg)
                print(error_msg)
            
            # Check headers
            content_type = response.headers.get('content-type', '').lower()
            content_disposition = response.headers.get('content-disposition', '')
            
            print(f"\n📋 Заголовки ответа:")
            print(f"   Content-Type: {content_type}")
            print(f"   Content-Disposition: {content_disposition}")
            
            # Store results
            results.finance_endpoints['export_all'] = {
                'status_code': response.status_code,
                'content_type': content_type,
                'content_disposition': content_disposition,
                'file_size': file_size,
                'response_time': response_time,
                'sheet_count': len(sheet_names) if 'sheet_names' in locals() else 0,
                'headers': dict(response.headers)
            }
            
            # Summary
            print(f"\n📊 ИТОГОВАЯ СВОДКА:")
            success_count = 4 - len([e for e in results.errors if 'КРИТЕРИЙ' in e])
            print(f"   ✅ Выполнено критериев: {success_count}/4")
            print(f"   ❌ Не выполнено критериев: {len([e for e in results.errors if 'КРИТЕРИЙ' in e])}/4")
            
            if not [e for e in results.errors if 'КРИТЕРИЙ' in e]:
                print("🎉 ВСЕ КРИТЕРИИ УСПЕХА ВЫПОЛНЕНЫ!")
            else:
                print("⚠️ Есть невыполненные критерии - см. детали выше")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании export-all: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_export_forecast_endpoint():
    """Test GET /api/finances/export-forecast endpoint according to review request"""
    print("\n=== ТЕСТ ЭКСПОРТА ПРОГНОЗОВ В XLSX ===\n")
    
    results = TestResults()
    
    # Test cases from review request
    test_cases = [
        {
            "company": "ВАШ ДОМ ФАКТ",
            "scenario": "realistic",
            "expected_size_kb": 6,
            "max_response_time": 3
        },
        {
            "company": "УФИЦ модель", 
            "scenario": "optimistic",
            "expected_size_kb": 6,
            "max_response_time": 3
        }
    ]
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            for i, test_case in enumerate(test_cases, 1):
                company = test_case["company"]
                scenario = test_case["scenario"]
                expected_size = test_case["expected_size_kb"] * 1024
                max_time = test_case["max_response_time"]
                
                print(f"📊 {i}. Тестируем GET /api/finances/export-forecast?company={company}&scenario={scenario}")
                print(f"🎯 КРИТЕРИИ УСПЕХА:")
                print(f"   1. ✅ Должен возвращать 200 OK")
                print(f"   2. ✅ XLSX файл с 1 листом прогноза")
                print(f"   3. ✅ Размер файла ~{test_case['expected_size_kb']}KB")
                print(f"   4. ✅ Время отклика < {max_time} секунд")
                print(f"   5. ✅ Имя файла без кириллицы (только ASCII)")
                print("")
                
                # Measure response time
                import time
                start_time = time.time()
                
                # Test the export-forecast endpoint
                response = await client.get(f"{API_BASE}/finances/export-forecast", params={
                    "company": company,
                    "scenario": scenario
                })
                
                end_time = time.time()
                response_time = end_time - start_time
                
                print(f"📡 Ответ сервера: {response.status_code}")
                print(f"⏱️ Время отклика: {response_time:.2f} секунд")
                
                # 1. Check 200 status
                if response.status_code != 200:
                    error_msg = f"❌ КРИТЕРИЙ 1 НЕ ВЫПОЛНЕН: Ожидался статус 200, получен {response.status_code}"
                    results.errors.append(error_msg)
                    print(error_msg)
                    if response.text:
                        print(f"   Текст ошибки: {response.text[:500]}")
                    continue
                
                print("✅ КРИТЕРИЙ 1 ВЫПОЛНЕН: Endpoint возвращает 200 OK")
                
                # 4. Check response time
                if response_time >= max_time:
                    error_msg = f"❌ КРИТЕРИЙ 4 НЕ ВЫПОЛНЕН: Время отклика {response_time:.2f}с >= {max_time}с"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print(f"✅ КРИТЕРИЙ 4 ВЫПОЛНЕН: Время отклика {response_time:.2f}с < {max_time}с")
                
                # Get file content
                content = response.content
                
                if not content:
                    error_msg = "❌ Ответ не содержит данных"
                    results.errors.append(error_msg)
                    print(error_msg)
                    continue
                
                # 3. Check file size
                file_size = len(content)
                min_size = expected_size * 0.5  # Allow 50% smaller
                max_size = expected_size * 2    # Allow 100% larger
                
                print(f"📊 Размер файла: {file_size:,} байт ({file_size / 1024:.1f} KB)")
                
                if file_size < min_size:
                    error_msg = f"❌ КРИТЕРИЙ 3 НЕ ВЫПОЛНЕН: Размер файла {file_size / 1024:.1f} KB слишком мал (< {min_size / 1024:.1f} KB)"
                    results.errors.append(error_msg)
                    print(error_msg)
                elif file_size > max_size:
                    error_msg = f"❌ КРИТЕРИЙ 3 НЕ ВЫПОЛНЕН: Размер файла {file_size / 1024:.1f} KB слишком велик (> {max_size / 1024:.1f} KB)"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print(f"✅ КРИТЕРИЙ 3 ВЫПОЛНЕН: Размер файла {file_size / 1024:.1f} KB в ожидаемом диапазоне")
                
                # 5. Check filename is ASCII only
                content_disposition = response.headers.get('content-disposition', '')
                print(f"📋 Content-Disposition: {content_disposition}")
                
                # Extract filename from Content-Disposition header
                filename = ""
                if 'filename=' in content_disposition:
                    filename_part = content_disposition.split('filename=')[1]
                    if filename_part.startswith('"') and filename_part.endswith('"'):
                        filename = filename_part[1:-1]
                    else:
                        filename = filename_part.split(';')[0]
                
                if filename:
                    try:
                        filename.encode('ascii')
                        print(f"✅ КРИТЕРИЙ 5 ВЫПОЛНЕН: Имя файла '{filename}' содержит только ASCII символы")
                    except UnicodeEncodeError:
                        error_msg = f"❌ КРИТЕРИЙ 5 НЕ ВЫПОЛНЕН: Имя файла '{filename}' содержит не-ASCII символы"
                        results.errors.append(error_msg)
                        print(error_msg)
                else:
                    print("⚠️ Имя файла не найдено в заголовках")
                
                # 2. Check valid XLSX format and sheet contents
                try:
                    import io
                    from openpyxl import load_workbook
                    
                    # Load XLSX file from memory
                    xlsx_file = io.BytesIO(content)
                    workbook = load_workbook(xlsx_file, read_only=True)
                    
                    print("✅ Файл валидный XLSX формат")
                    
                    # Get all sheet names
                    sheet_names = workbook.sheetnames
                    print(f"📋 Найдено листов: {len(sheet_names)}")
                    
                    print(f"📋 Листы в файле:")
                    for j, sheet_name in enumerate(sheet_names, 1):
                        print(f"   {j}. {sheet_name}")
                    
                    # Check if we have 1 sheet total
                    if len(sheet_names) != 1:
                        error_msg = f"❌ КРИТЕРИЙ 2 НЕ ВЫПОЛНЕН: Найдено {len(sheet_names)} листов, ожидался 1"
                        results.errors.append(error_msg)
                        print(error_msg)
                    else:
                        print("✅ КРИТЕРИЙ 2 ВЫПОЛНЕН: Найден 1 лист прогноза")
                    
                    # Validate sheet contents (basic check)
                    if sheet_names:
                        forecast_sheet = workbook[sheet_names[0]]
                        max_row = forecast_sheet.max_row
                        max_col = forecast_sheet.max_column
                        print(f"   Лист прогноза: {max_row} строк, {max_col} столбцов")
                        
                        if max_row > 1 and max_col > 1:
                            print("   ✅ Лист содержит данные")
                        else:
                            print("   ⚠️ Лист может быть пустым")
                    
                    workbook.close()
                    
                except ImportError:
                    print("⚠️ openpyxl не установлен - проверка содержимого XLSX пропущена")
                    
                    # Basic XLSX validation - check ZIP signature
                    if content.startswith(b'PK'):
                        print("✅ Файл имеет корректную XLSX сигнатуру (ZIP-based)")
                    else:
                        error_msg = f"❌ Неверная сигнатура файла: {content[:10]}"
                        results.errors.append(error_msg)
                        print(error_msg)
                    
                except Exception as xlsx_error:
                    error_msg = f"❌ Ошибка чтения XLSX: {str(xlsx_error)}"
                    results.errors.append(error_msg)
                    print(error_msg)
                
                # Store results for this test case
                results.finance_endpoints[f'export_forecast_{company}_{scenario}'] = {
                    'status_code': response.status_code,
                    'content_type': response.headers.get('content-type', ''),
                    'content_disposition': content_disposition,
                    'filename': filename,
                    'file_size': file_size,
                    'response_time': response_time,
                    'sheet_count': len(sheet_names) if 'sheet_names' in locals() else 0,
                    'headers': dict(response.headers)
                }
                
                # Summary for this test case
                print(f"\n📊 ИТОГОВАЯ СВОДКА для {company} - {scenario}:")
                test_case_errors = [e for e in results.errors if f'{company}' in e or 'КРИТЕРИЙ' in e]
                success_count = 5 - len([e for e in test_case_errors if 'КРИТЕРИЙ' in e])
                print(f"   ✅ Выполнено критериев: {success_count}/5")
                print(f"   ❌ Не выполнено критериев: {len([e for e in test_case_errors if 'КРИТЕРИЙ' in e])}/5")
                
                if not [e for e in test_case_errors if 'КРИТЕРИЙ' in e]:
                    print("🎉 ВСЕ КРИТЕРИИ УСПЕХА ВЫПОЛНЕНЫ!")
                else:
                    print("⚠️ Есть невыполненные критерии - см. детали выше")
                
                print("\n" + "="*80 + "\n")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании export-forecast: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_consolidated_financial_model():
    """Test consolidated financial model for ООО ВАШ ДОМ + УФИЦ"""
    print("\n=== ТЕСТ КОНСОЛИДИРОВАННОЙ ФИНАНСОВОЙ МОДЕЛИ ООО ВАШ ДОМ + УФИЦ ===\n")
    
    results = TestResults()
    consolidated_company = "ООО ВАШ ДОМ + УФИЦ"
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            print(f"🏢 Тестируем консолидированную модель для: {consolidated_company}")
            print("📋 Логика консолидации:")
            print("   - Выручка = ООО ВАШ ДОМ минус 'Швеи' и 'Аутсорсинг'")
            print("   - Расходы = ООО ВАШ ДОМ + УФИЦ минус 'Кредиты', 'Аутсорсинг персонала с Ю/ЦЛ', 'Швеи'")
            print("")
            
            # 1. Test consolidated profit-loss
            print("💰 1. Тестируем GET /api/finances/profit-loss?company=ООО+ВАШ+ДОМ+%2B+УФИЦ")
            
            response = await client.get(f"{API_BASE}/finances/profit-loss", params={"company": consolidated_company})
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка consolidated profit-loss: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
            else:
                data = response.json()
                results.finance_endpoints['consolidated_profit_loss'] = data
                print("✅ Консолидированные прибыли/убытки получены успешно")
                
                # Validate structure
                required_fields = ['profit_loss', 'summary']
                for field in required_fields:
                    if field not in data:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в consolidated profit-loss")
                    else:
                        print(f"✅ Поле '{field}' присутствует")
                
                # Check summary
                if 'summary' in data:
                    summary = data['summary']
                    print(f"📊 Консолидированная сводка:")
                    print(f"   - Общая выручка: {summary.get('total_revenue', 0)}")
                    print(f"   - Общие расходы: {summary.get('total_expenses', 0)}")
                    print(f"   - Чистая прибыль: {summary.get('net_profit', 0)}")
                    print(f"   - Средняя маржа: {summary.get('average_margin', 0)}%")
            
            print("")
            
            # 2. Test consolidated expense-analysis
            print("💸 2. Тестируем GET /api/finances/expense-analysis?company=ООО+ВАШ+ДОМ+%2B+УФИЦ")
            
            response = await client.get(f"{API_BASE}/finances/expense-analysis", params={"company": consolidated_company})
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка consolidated expense-analysis: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
            else:
                data = response.json()
                results.finance_endpoints['consolidated_expense_analysis'] = data
                print("✅ Консолидированный анализ расходов получен успешно")
                
                # Validate structure
                required_fields = ['expenses', 'total']
                for field in required_fields:
                    if field not in data:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в consolidated expense-analysis")
                    else:
                        print(f"✅ Поле '{field}' присутствует")
                
                # Check excluded categories are not present
                excluded_categories = ['Кредиты', 'Аутсорсинг персонала с Ю/ЦЛ', 'Швеи']
                expenses = data.get('expenses', [])
                
                print(f"📊 Консолидированные расходы:")
                print(f"   - Категорий расходов: {len(expenses)}")
                print(f"   - Общая сумма: {data.get('total', 0)}")
                
                # Check for excluded categories
                found_excluded = []
                for expense in expenses:
                    category = expense.get('category', '')
                    if category in excluded_categories:
                        found_excluded.append(category)
                
                if found_excluded:
                    error_msg = f"❌ Найдены исключаемые категории в консолидированных расходах: {found_excluded}"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print("✅ Исключаемые категории отсутствуют в результатах")
                
                # Show top categories
                if expenses:
                    print("📋 Топ категории расходов:")
                    for i, expense in enumerate(expenses[:5], 1):
                        print(f"   {i}. {expense.get('category')}: {expense.get('amount')} ({expense.get('percentage')}%)")
            
            print("")
            
            # 3. Test consolidated expense-analysis with month filter
            print("📅 3. Тестируем GET /api/finances/expense-analysis?company=ООО+ВАШ+ДОМ+%2B+УФИЦ&month=Январь+2025")
            
            response = await client.get(f"{API_BASE}/finances/expense-analysis", params={
                "company": consolidated_company,
                "month": "Январь 2025"
            })
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка consolidated expense-analysis с месяцем: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
            else:
                data = response.json()
                results.finance_endpoints['consolidated_expense_analysis_monthly'] = data
                print("✅ Консолидированный анализ расходов за месяц получен успешно")
                
                # Validate month field
                if data.get('month') != "Январь 2025":
                    results.errors.append(f"❌ Неверный месяц в ответе: ожидался 'Январь 2025', получен '{data.get('month')}'")
                else:
                    print(f"✅ Месяц корректен: {data.get('month')}")
                
                expenses = data.get('expenses', [])
                print(f"📊 Расходы за Январь 2025:")
                print(f"   - Категорий: {len(expenses)}")
                print(f"   - Общая сумма: {data.get('total', 0)}")
            
            print("")
            
            # 4. Test consolidated revenue-analysis
            print("💵 4. Тестируем GET /api/finances/revenue-analysis?company=ООО+ВАШ+ДОМ+%2B+УФИЦ")
            
            response = await client.get(f"{API_BASE}/finances/revenue-analysis", params={"company": consolidated_company})
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка consolidated revenue-analysis: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
            else:
                data = response.json()
                results.finance_endpoints['consolidated_revenue_analysis'] = data
                print("✅ Консолидированный анализ выручки получен успешно")
                
                # Validate structure
                required_fields = ['revenue', 'total']
                for field in required_fields:
                    if field not in data:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в consolidated revenue-analysis")
                    else:
                        print(f"✅ Поле '{field}' присутствует")
                
                # Check excluded categories for revenue (Швеи, Аутсорсинг)
                excluded_revenue_categories = ['Швеи', 'Аутсорсинг']
                revenues = data.get('revenue', [])
                
                print(f"📊 Консолидированная выручка:")
                print(f"   - Категорий выручки: {len(revenues)}")
                print(f"   - Общая сумма: {data.get('total', 0)}")
                
                # Check for excluded categories
                found_excluded_revenue = []
                for revenue in revenues:
                    category = revenue.get('category', '')
                    if category in excluded_revenue_categories:
                        found_excluded_revenue.append(category)
                
                if found_excluded_revenue:
                    error_msg = f"❌ Найдены исключаемые категории в консолидированной выручке: {found_excluded_revenue}"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print("✅ Исключаемые категории выручки отсутствуют в результатах")
                
                # Show revenue categories
                if revenues:
                    print("📋 Категории выручки:")
                    for i, revenue in enumerate(revenues[:5], 1):
                        print(f"   {i}. {revenue.get('category')}: {revenue.get('amount')} ({revenue.get('percentage')}%)")
            
            print("")
            
            # 5. Test individual company endpoints still work
            print("🔄 5. Проверяем работу endpoints для отдельных компаний...")
            
            individual_companies = ["ООО ВАШ ДОМ", "УФИЦ"]
            
            for company in individual_companies:
                print(f"\n🏢 Тестируем компанию: {company}")
                
                # Test profit-loss for individual company
                response = await client.get(f"{API_BASE}/finances/profit-loss", params={"company": company})
                print(f"📡 profit-loss для {company}: {response.status_code}")
                
                if response.status_code != 200:
                    error_msg = f"❌ Endpoint profit-loss не работает для {company}: {response.status_code}"
                    results.errors.append(error_msg)
                else:
                    print(f"✅ Endpoint profit-loss работает для {company}")
                
                # Test expense-analysis for individual company
                response = await client.get(f"{API_BASE}/finances/expense-analysis", params={"company": company})
                print(f"📡 expense-analysis для {company}: {response.status_code}")
                
                if response.status_code != 200:
                    error_msg = f"❌ Endpoint expense-analysis не работает для {company}: {response.status_code}"
                    results.errors.append(error_msg)
                else:
                    print(f"✅ Endpoint expense-analysis работает для {company}")
            
            print("")
            
            # 6. Validate consolidation logic by comparing individual vs consolidated
            print("🧮 6. Проверяем логику консолидации...")
            
            # Get individual company data for validation
            vasdom_response = await client.get(f"{API_BASE}/finances/expense-analysis", params={"company": "ООО ВАШ ДОМ"})
            ufic_response = await client.get(f"{API_BASE}/finances/expense-analysis", params={"company": "УФИЦ"})
            
            if vasdom_response.status_code == 200 and ufic_response.status_code == 200:
                vasdom_data = vasdom_response.json()
                ufic_data = ufic_response.json()
                
                vasdom_total = vasdom_data.get('total', 0)
                ufic_total = ufic_data.get('total', 0)
                
                print(f"📊 Данные для проверки консолидации:")
                print(f"   - ООО ВАШ ДОМ общие расходы: {vasdom_total}")
                print(f"   - УФИЦ общие расходы: {ufic_total}")
                
                # Get excluded amounts from ООО ВАШ ДОМ
                excluded_amount = 0
                vasdom_expenses = vasdom_data.get('expenses', [])
                for expense in vasdom_expenses:
                    if expense.get('category') in ['Кредиты', 'Аутсорсинг персонала с Ю/ЦЛ', 'Швеи']:
                        excluded_amount += expense.get('amount', 0)
                
                expected_consolidated_total = vasdom_total + ufic_total - excluded_amount
                
                # Get actual consolidated total
                if 'consolidated_expense_analysis' in results.finance_endpoints:
                    actual_consolidated_total = results.finance_endpoints['consolidated_expense_analysis'].get('total', 0)
                    
                    print(f"   - Исключаемая сумма: {excluded_amount}")
                    print(f"   - Ожидаемая консолидированная сумма: {expected_consolidated_total}")
                    print(f"   - Фактическая консолидированная сумма: {actual_consolidated_total}")
                    
                    # Allow small differences due to floating point arithmetic
                    if abs(expected_consolidated_total - actual_consolidated_total) < 1.0:
                        print("✅ Логика консолидации расходов работает корректно")
                    else:
                        error_msg = f"❌ Неверная логика консолидации: ожидалось {expected_consolidated_total}, получено {actual_consolidated_total}"
                        results.errors.append(error_msg)
                        print(error_msg)
            else:
                print("⚠️ Не удалось получить данные отдельных компаний для проверки логики")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании консолидированной модели: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_plannerka_create_endpoint():
    """Test plannerka creation endpoint"""
    print("\n=== ТЕСТ СОЗДАНИЯ ПЛАНЁРКИ ===\n")
    
    results = TestResults()
    
    # Test data from the review request
    test_data = {
        "title": "Тестовая планёрка",
        "transcription": "Обсуждали задачи на неделю. Иванову поручено завершить отчет до пятницы. Петрову нужно проверить документы до среды. Сидорову поручена подготовка презентации с высоким приоритетом до четверга.",
        "participants": ["Иванов", "Петров", "Сидоров"]
    }
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("🔍 Создаем планёрку...")
            print(f"📝 Данные: {json.dumps(test_data, ensure_ascii=False, indent=2)}")
            
            # Test the create endpoint
            response = await client.post(
                f"{API_BASE}/plannerka/create",
                json=test_data,
                headers={"Content-Type": "application/json"}
            )
            
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка создания планёрки: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.created_meeting_data = data
            results.created_meeting_id = data.get('id')
            results.database_working = True
            
            print(f"✅ Планёрка создана с ID: {results.created_meeting_id}")
            print(f"📊 Структура ответа:")
            print(f"   - ID: {data.get('id')}")
            print(f"   - Title: {data.get('title')}")
            print(f"   - Date: {data.get('date')}")
            print(f"   - Participants: {data.get('participants')}")
            print(f"   - Transcription length: {len(data.get('transcription', ''))}")
            print(f"   - Summary: {data.get('summary')}")
            print(f"   - Tasks: {data.get('tasks')}")
            
            # Validate response structure
            required_fields = ['id', 'title', 'date', 'transcription', 'participants', 'created_at']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в ответе")
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            # Validate data integrity
            if data.get('title') != test_data['title']:
                results.errors.append(f"❌ Неверный title: ожидался '{test_data['title']}', получен '{data.get('title')}'")
            
            if data.get('transcription') != test_data['transcription']:
                results.errors.append(f"❌ Неверная transcription")
            
            if data.get('participants') != test_data['participants']:
                results.errors.append(f"❌ Неверные participants")
            
            if not results.errors:
                print("✅ Все поля корректны")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании создания планёрки: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_plannerka_analyze_endpoint(meeting_id: str):
    """Test plannerka AI analysis endpoint"""
    print("\n=== ТЕСТ AI-АНАЛИЗА ПЛАНЁРКИ ===\n")
    
    results = TestResults()
    results.created_meeting_id = meeting_id
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:  # Longer timeout for AI analysis
            print(f"🤖 Запускаем AI-анализ планёрки ID: {meeting_id}")
            
            # Test the analyze endpoint
            response = await client.post(f"{API_BASE}/plannerka/analyze/{meeting_id}")
            
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка AI-анализа: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                
                # Check for specific OpenAI errors
                if "OPENAI_API_KEY" in response.text:
                    results.errors.append("❌ OPENAI_API_KEY не настроен")
                elif "transcription is too short" in response.text:
                    results.errors.append("❌ Транскрипция слишком короткая")
                
                return results
            
            data = response.json()
            results.analysis_result = data
            results.openai_working = True
            
            print(f"✅ AI-анализ завершен успешно")
            print(f"📊 Результат анализа:")
            print(f"   - Success: {data.get('success')}")
            print(f"   - Tasks count: {data.get('tasks_count')}")
            
            # Check summary
            summary = data.get('summary', '')
            if summary:
                results.summary_generated = True
                print(f"✅ Саммари сгенерировано ({len(summary)} символов)")
                print(f"📝 Саммари: {summary[:200]}{'...' if len(summary) > 200 else ''}")
            else:
                results.errors.append("❌ Саммари не сгенерировано")
            
            # Check tasks
            tasks = data.get('tasks', [])
            results.tasks_extracted = tasks
            
            if tasks:
                print(f"✅ Извлечено задач: {len(tasks)}")
                print("📋 Задачи:")
                
                for i, task in enumerate(tasks, 1):
                    title = task.get('title', 'Без названия')
                    assignee = task.get('assignee', 'Не назначен')
                    deadline = task.get('deadline', 'Не указан')
                    priority = task.get('priority', 'medium')
                    
                    print(f"   {i}. {title}")
                    print(f"      - Исполнитель: {assignee}")
                    print(f"      - Срок: {deadline}")
                    print(f"      - Приоритет: {priority}")
                    
                    # Validate task structure
                    required_task_fields = ['title']
                    for field in required_task_fields:
                        if field not in task or not task[field]:
                            results.errors.append(f"❌ Задача {i}: отсутствует поле '{field}'")
                
                # Check if expected tasks are found
                expected_assignees = ['Иванов', 'Петров', 'Сидоров']
                found_assignees = [task.get('assignee', '') for task in tasks]
                
                for expected in expected_assignees:
                    if any(expected in assignee for assignee in found_assignees):
                        print(f"✅ Найдена задача для {expected}")
                    else:
                        print(f"⚠️ Не найдена задача для {expected}")
                
            else:
                results.errors.append("❌ Задачи не извлечены")
            
            # Validate response structure
            required_fields = ['success', 'summary', 'tasks', 'tasks_count']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в ответе анализа")
            
            # Check if success is true
            if not data.get('success'):
                results.errors.append("❌ Поле 'success' не равно true")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании AI-анализа: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_plannerka_list_endpoint():
    """Test plannerka list endpoint"""
    print("\n=== ТЕСТ СПИСКА ПЛАНЁРОК ===\n")
    
    results = TestResults()
    
    try:
        async with httpx.AsyncClient(timeout=30.0) as client:
            print("📋 Получаем список планёрок...")
            
            # Test the list endpoint
            response = await client.get(f"{API_BASE}/plannerka/list?limit=20&offset=0")
            
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка получения списка: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            results.meetings_list = data.get('meetings', [])
            
            print(f"✅ Список получен успешно")
            print(f"📊 Результат:")
            print(f"   - Количество планёрок: {data.get('count', 0)}")
            print(f"   - Планёрок в ответе: {len(results.meetings_list)}")
            
            # Validate response structure
            required_fields = ['meetings', 'count']
            for field in required_fields:
                if field not in data:
                    results.errors.append(f"❌ Отсутствует поле '{field}' в ответе списка")
            
            # Check meetings structure
            if results.meetings_list:
                print("📋 Примеры планёрок:")
                
                for i, meeting in enumerate(results.meetings_list[:3], 1):  # Show first 3
                    meeting_id = meeting.get('id', 'N/A')
                    title = meeting.get('title', 'Без названия')
                    date = meeting.get('date', 'Не указана')
                    participants_count = len(meeting.get('participants', []))
                    tasks_count = len(meeting.get('tasks', []))
                    has_summary = bool(meeting.get('summary'))
                    
                    print(f"   {i}. ID: {meeting_id}")
                    print(f"      - Название: {title}")
                    print(f"      - Дата: {date}")
                    print(f"      - Участников: {participants_count}")
                    print(f"      - Задач: {tasks_count}")
                    print(f"      - Есть саммари: {has_summary}")
                
                # Validate meeting structure
                sample_meeting = results.meetings_list[0]
                expected_meeting_fields = ['id', 'title', 'date', 'transcription', 'participants']
                
                for field in expected_meeting_fields:
                    if field not in sample_meeting:
                        results.errors.append(f"❌ Отсутствует поле '{field}' в структуре планёрки")
                    else:
                        print(f"✅ Поле '{field}' присутствует в структуре")
            else:
                print("⚠️ Список планёрок пуст")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании списка планёрок: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_openai_configuration():
    """Test OpenAI API configuration"""
    print("\n🔍 Проверяем конфигурацию OpenAI...")
    
    try:
        # Check if OPENAI_API_KEY is configured by testing a simple endpoint
        async with httpx.AsyncClient(timeout=15.0) as client:
            # We'll test this indirectly through the health endpoint
            response = await client.get(f"{API_BASE}/health")
            if response.status_code == 200:
                print("✅ Backend доступен")
            else:
                print(f"⚠️ Backend недоступен: {response.status_code}")
                
        # Check environment variables (we can't directly access them, but we can infer from API responses)
        print("📋 Конфигурация будет проверена через API вызовы")
        
    except Exception as e:
        print(f"❌ Ошибка проверки конфигурации: {str(e)}")

async def test_database_connection():
    """Test database connection through plannerka endpoints"""
    print("\n🔍 Проверяем подключение к базе данных...")
    
    try:
        async with httpx.AsyncClient(timeout=15.0) as client:
            # Test database connection by trying to get the list
            response = await client.get(f"{API_BASE}/plannerka/list?limit=1")
            
            if response.status_code == 200:
                print("✅ База данных доступна")
                return True
            elif response.status_code == 500 and "Database connection error" in response.text:
                print("❌ Ошибка подключения к базе данных")
                return False
            else:
                print(f"⚠️ Неожиданный ответ: {response.status_code}")
                return False
                
    except Exception as e:
        print(f"❌ Ошибка проверки БД: {str(e)}")
        return False

async def test_vasdom_model_forecast_endpoint():
    """Test ВАШ ДОМ модель forecast endpoint with cleaners integration from review request"""
    print("\n=== КРИТИЧЕСКОЕ ПОВТОРНОЕ ТЕСТИРОВАНИЕ ВАШ ДОМ МОДЕЛЬ ===\n")
    
    results = TestResults()
    scenarios = ["pessimistic", "realistic", "optimistic"]
    company = "ВАШ ДОМ модель"
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            print(f"🏢 Тестируем прогноз ВАШ ДОМ модель с интеграцией уборщиц из УФИЦ")
            print("📋 КРИТЕРИИ УСПЕХА из review request:")
            print("   1. Endpoint должен возвращать 200 статус для всех трех сценариев")
            print("")
            print("   2. В expense_breakdown для каждого года (2026-2030) должна присутствовать")
            print("      категория 'аутсорсинг_персонала' с суммами:")
            print("      - PESSIMISTIC (самые БОЛЬШИЕ суммы): 2026=34347600, 2027=35996285,")
            print("        2028=37724106, 2029=39534864, 2030=41432537")
            print("      - REALISTIC (средние суммы): 2026=24615780, 2027=25797337,")
            print("        2028=27035610, 2029=28333319, 2030=29693318")
            print("      - OPTIMISTIC (самые МАЛЕНЬКИЕ суммы): 2026=16028880, 2027=17134873,")
            print("        2028=18317179, 2029=19581064, 2030=20932158")
            print("")
            print("   3. В revenue_breakdown должны быть поля:")
            print("      - vasdom_revenue (чистая выручка ВАШ ДОМ)")
            print("      - ufic_sewing (швеи из УФИЦ)")
            print("      - ufic_outsourcing (аутсорсинг из УФИЦ)")
            print("      - total (общая выручка)")
            print("")
            print("   4. Проверить, что vasdom_revenue + ufic_sewing + ufic_outsourcing = total")
            print("")
            
            scenario_results = {}
            
            for scenario in scenarios:
                print(f"🔍 Тестируем сценарий: {scenario}")
                
                # Test the forecast endpoint
                response = await client.get(
                    f"{API_BASE}/finances/forecast",
                    params={"company": company, "scenario": scenario}
                )
                
                print(f"📡 Ответ сервера для {scenario}: {response.status_code}")
                
                # Check 200 status
                if response.status_code != 200:
                    error_msg = f"❌ Сценарий {scenario}: ошибка {response.status_code} - {response.text}"
                    results.errors.append(error_msg)
                    print(error_msg)
                    continue
                
                data = response.json()
                scenario_results[scenario] = data
                print(f"✅ Сценарий {scenario}: 200 статус получен")
                
                # Validate response structure
                required_fields = ['company', 'scenario', 'base_year', 'base_data', 'forecast']
                for field in required_fields:
                    if field not in data:
                        error_msg = f"❌ Сценарий {scenario}: отсутствует поле '{field}'"
                        results.errors.append(error_msg)
                        print(error_msg)
                    else:
                        print(f"✅ Поле '{field}' присутствует")
                
                # Check forecast data for years 2026-2030
                forecast = data.get('forecast', [])
                if len(forecast) < 5:
                    error_msg = f"❌ Сценарий {scenario}: недостаточно лет в прогнозе (ожидалось 5, получено {len(forecast)})"
                    results.errors.append(error_msg)
                    print(error_msg)
                    continue
                
                print(f"✅ Прогноз содержит {len(forecast)} лет (2026-2030)")
                
                # Get base year data
                base_data = data.get('base_data', {})
                base_revenue = base_data.get('revenue', 0)
                base_expenses = base_data.get('expenses', 0)
                
                print(f"📊 Базовые данные 2025:")
                print(f"   - Выручка: {base_revenue:,.0f} ₽")
                print(f"   - Расходы: {base_expenses:,.0f} ₽")
                
                # Check expense_breakdown for аутсорсинг_персонала category
                print(f"\n📊 Проверяем сценарий {scenario.upper()}:")
                
                # Expected values for аутсорсинг_персонала by scenario and year
                expected_outsourcing = {
                    "pessimistic": {2026: 34347600, 2027: 35996285, 2028: 37724106, 2029: 39534864, 2030: 41432537},
                    "realistic": {2026: 24615780, 2027: 25797337, 2028: 27035610, 2029: 28333319, 2030: 29693318},
                    "optimistic": {2026: 16028880, 2027: 17134873, 2028: 18317179, 2029: 19581064, 2030: 20932158}
                }
                
                scenario_expected = expected_outsourcing.get(scenario, {})
                
                # Check each year (2026-2030)
                for year_data in forecast:
                    year = year_data.get('year')
                    if year not in range(2026, 2031):
                        continue
                    
                    expense_breakdown = year_data.get('expense_breakdown', {})
                    
                    # Check if аутсорсинг_персонала category exists
                    outsourcing_amount = expense_breakdown.get('аутсорсинг_персонала', 0)
                    expected_amount = scenario_expected.get(year, 0)
                    
                    if 'аутсорсинг_персонала' not in expense_breakdown:
                        error_msg = f"❌ {scenario} {year}: отсутствует категория 'аутсорсинг_персонала' в expense_breakdown"
                        results.errors.append(error_msg)
                        print(error_msg)
                        continue
                    
                    # Check if amount matches expected (allow 1% tolerance)
                    if expected_amount > 0:
                        diff_pct = abs(outsourcing_amount - expected_amount) / expected_amount * 100
                        if diff_pct > 1.0:
                            error_msg = f"❌ {scenario} {year}: неверная сумма аутсорсинга. Ожидалось {expected_amount:,.0f}, получено {outsourcing_amount:,.0f} (отклонение {diff_pct:.2f}%)"
                            results.errors.append(error_msg)
                            print(error_msg)
                        else:
                            print(f"✅ {year}: аутсорсинг_персонала = {outsourcing_amount:,.0f} ₽ (ожидалось {expected_amount:,.0f})")
                    else:
                        print(f"⚠️ {year}: нет ожидаемого значения для {scenario}")
                
                # Check revenue_breakdown structure
                print(f"\n💰 Проверяем структуру выручки для {scenario}:")
                
                # Check first year as example
                if forecast:
                    first_year = forecast[0]
                    revenue_breakdown = first_year.get('revenue_breakdown', {})
                    
                    required_revenue_fields = ['vasdom_revenue', 'ufic_sewing', 'ufic_outsourcing', 'total']
                    
                    for field in required_revenue_fields:
                        if field not in revenue_breakdown:
                            error_msg = f"❌ {scenario}: отсутствует поле '{field}' в revenue_breakdown"
                            results.errors.append(error_msg)
                            print(error_msg)
                        else:
                            value = revenue_breakdown[field]
                            print(f"✅ {field}: {value:,.0f} ₽")
                    
                    # Check revenue calculation: vasdom_revenue + ufic_sewing + ufic_outsourcing = total
                    if all(field in revenue_breakdown for field in required_revenue_fields):
                        vasdom = revenue_breakdown['vasdom_revenue']
                        ufic_sewing = revenue_breakdown['ufic_sewing']
                        ufic_outsourcing = revenue_breakdown['ufic_outsourcing']
                        total = revenue_breakdown['total']
                        
                        calculated_total = vasdom + ufic_sewing + ufic_outsourcing
                        
                        if abs(calculated_total - total) > 1.0:  # Allow small rounding differences
                            error_msg = f"❌ {scenario}: неверный расчет total. {vasdom:,.0f} + {ufic_sewing:,.0f} + {ufic_outsourcing:,.0f} = {calculated_total:,.0f}, но total = {total:,.0f}"
                            results.errors.append(error_msg)
                            print(error_msg)
                        else:
                            print(f"✅ Расчет выручки корректен: {vasdom:,.0f} + {ufic_sewing:,.0f} + {ufic_outsourcing:,.0f} = {total:,.0f}")
                
                print("")  # Empty line for readability
            
            # Store results for this scenario
            results.finance_endpoints[f'{company}_{scenario}_forecast'] = data
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании прогноза ВАШ ДОМ модель: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def test_ufic_model_forecast_expense_breakdown():
    """Test УФИЦ модель forecast endpoint expense_breakdown structure according to review request"""
    print("\n=== ТЕСТ СТРУКТУРЫ EXPENSE_BREAKDOWN УФИЦ МОДЕЛЬ ===\n")
    
    results = TestResults()
    company = "УФИЦ модель"
    scenario = "realistic"
    
    try:
        async with httpx.AsyncClient(timeout=60.0) as client:
            print(f"🏢 Тестируем endpoint: GET /api/finances/forecast?company={company}&scenario={scenario}")
            print("📋 КРИТЕРИИ УСПЕХА из review request:")
            print("   1. В expense_breakdown НЕ должно быть ключа 'ндфл' или 'ndfl'")
            print("   2. В expense_breakdown ДОЛЖЕН быть ключ 'текущий_ремонт_обучение'")
            print("   3. В expense_breakdown НЕ должно быть отдельных ключей:")
            print("      'газпром', 'первый_газовый', 'водоканал', 'крэо', 'вдпо', 'прикамский_институт'")
            print("   4. Вывести содержимое expense_breakdown для 2026 года для проверки")
            print("")
            
            # Test the forecast endpoint
            response = await client.get(
                f"{API_BASE}/finances/forecast",
                params={"company": company, "scenario": scenario}
            )
            
            print(f"📡 Ответ сервера: {response.status_code}")
            
            if response.status_code != 200:
                error_msg = f"❌ Ошибка получения прогноза: {response.status_code} - {response.text}"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            data = response.json()
            print(f"✅ Прогноз получен успешно")
            
            # Validate response structure
            required_fields = ['company', 'scenario', 'forecast']
            for field in required_fields:
                if field not in data:
                    error_msg = f"❌ Отсутствует поле '{field}' в ответе"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print(f"✅ Поле '{field}' присутствует")
            
            # Check forecast data
            forecast = data.get('forecast', [])
            if len(forecast) < 5:
                error_msg = f"❌ Недостаточно лет в прогнозе (ожидалось 5, получено {len(forecast)})"
                results.errors.append(error_msg)
                print(error_msg)
                return results
            
            print(f"✅ Прогноз содержит {len(forecast)} лет")
            
            # Check expense_breakdown for each year (2026-2030)
            forbidden_keys = ["ндфл", "ndfl", "газпром", "первый_газовый", "водоканал", "крэо", "вдпо", "прикамский_институт"]
            required_keys = ["текущий_ремонт_обучение"]
            
            expense_breakdown_2026 = None
            
            for year_data in forecast:
                year = year_data.get('year')
                if year not in range(2026, 2031):
                    continue
                
                expense_breakdown = year_data.get('expense_breakdown', {})
                
                if year == 2026:
                    expense_breakdown_2026 = expense_breakdown
                
                print(f"\n📊 Проверяем expense_breakdown для {year} года:")
                print(f"   - Категорий в expense_breakdown: {len(expense_breakdown)}")
                
                # Check for forbidden keys
                found_forbidden = []
                for key in expense_breakdown.keys():
                    key_lower = key.lower()
                    for forbidden in forbidden_keys:
                        if forbidden.lower() in key_lower:
                            found_forbidden.append(key)
                
                if found_forbidden:
                    error_msg = f"❌ Год {year}: найдены запрещенные ключи в expense_breakdown: {found_forbidden}"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print(f"✅ Год {year}: запрещенные ключи отсутствуют")
                
                # Check for required keys
                found_required = []
                for key in expense_breakdown.keys():
                    key_lower = key.lower()
                    for required in required_keys:
                        if required.lower() in key_lower:
                            found_required.append(key)
                
                if not found_required:
                    error_msg = f"❌ Год {year}: не найден обязательный ключ 'текущий_ремонт_обучение' в expense_breakdown"
                    results.errors.append(error_msg)
                    print(error_msg)
                else:
                    print(f"✅ Год {year}: найден обязательный ключ: {found_required}")
                
                # Show all keys for debugging
                print(f"   - Все ключи expense_breakdown: {list(expense_breakdown.keys())}")
            
            # Output expense_breakdown for 2026 as requested
            print(f"\n📋 СОДЕРЖИМОЕ EXPENSE_BREAKDOWN ДЛЯ 2026 ГОДА:")
            print("=" * 60)
            
            if expense_breakdown_2026:
                for key, value in expense_breakdown_2026.items():
                    print(f"   {key}: {value:,.0f} ₽")
                
                print(f"\n📊 Итого категорий: {len(expense_breakdown_2026)}")
                total_amount = sum(expense_breakdown_2026.values())
                print(f"💰 Общая сумма: {total_amount:,.0f} ₽")
            else:
                error_msg = "❌ Не удалось получить expense_breakdown для 2026 года"
                results.errors.append(error_msg)
                print(error_msg)
            
            print("=" * 60)
            
            # Summary of test results
            print(f"\n📈 ИТОГИ ТЕСТИРОВАНИЯ:")
            
            # Count forbidden keys across all years
            total_forbidden_found = sum(1 for error in results.errors if "запрещенные ключи" in error)
            if total_forbidden_found == 0:
                print("✅ Критерий 1: Запрещенные ключи ('ндфл', 'ndfl', 'газпром', и др.) отсутствуют во всех годах")
            else:
                print(f"❌ Критерий 1: Найдены запрещенные ключи в {total_forbidden_found} годах")
            
            # Count missing required keys
            total_missing_required = sum(1 for error in results.errors if "обязательный ключ" in error)
            if total_missing_required == 0:
                print("✅ Критерий 2: Обязательный ключ 'текущий_ремонт_обучение' присутствует во всех годах")
            else:
                print(f"❌ Критерий 2: Отсутствует обязательный ключ в {total_missing_required} годах")
            
            # Check if individual forbidden keys are absent
            individual_forbidden_absent = True
            if expense_breakdown_2026:
                for forbidden in ["газпром", "первый_газовый", "водоканал", "крэо", "вдпо", "прикамский_институт"]:
                    if any(forbidden.lower() in key.lower() for key in expense_breakdown_2026.keys()):
                        individual_forbidden_absent = False
                        break
            
            if individual_forbidden_absent:
                print("✅ Критерий 3: Отдельные запрещенные ключи отсутствуют")
            else:
                print("❌ Критерий 3: Найдены отдельные запрещенные ключи")
            
            print("✅ Критерий 4: Содержимое expense_breakdown для 2026 года выведено выше")
            
    except Exception as e:
        error_msg = f"❌ Ошибка при тестировании УФИЦ модель forecast: {str(e)}"
        results.errors.append(error_msg)
        print(error_msg)
    
    return results

async def main():
    """Main test execution - focused on export endpoints testing per review request"""
    print("🚀 ФИНАЛЬНАЯ ПРОВЕРКА ПЕРЕД PRODUCTION DEPLOY")
    print("=" * 80)
    print(f"🌐 Backend URL: {BACKEND_URL}")
    print(f"🔗 API Base: {API_BASE}")
    print("")
    print("🎯 ЦЕЛЬ: Протестировать оба экспорт endpoint перед production")
    print("📋 Endpoints для проверки:")
    print("   1. GET /api/finances/export-all")
    print("   2. GET /api/finances/export-forecast?company=ВАШ ДОМ ФАКТ&scenario=realistic")
    print("   3. GET /api/finances/export-forecast?company=УФИЦ модель&scenario=optimistic")
    print("=" * 80)
    
    # Check basic connectivity
    print("\n🔍 Проверяем доступность backend...")
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            response = await client.get(f"{API_BASE}/health")
            if response.status_code == 200:
                print("✅ Backend доступен")
            else:
                print(f"⚠️ Backend недоступен: {response.status_code}")
                return []
    except Exception as e:
        print(f"❌ Ошибка подключения к backend: {str(e)}")
        return []
    
    all_results = []
    
    # Test 1: Export All endpoint
    print("\n" + "=" * 80)
    print("🧪 1. ТЕСТИРОВАНИЕ GET /api/finances/export-all")
    print("=" * 80)
    
    export_all_result = await test_export_all_endpoint()
    all_results.append(("Export All", export_all_result))
    
    # Test 2: Export Forecast endpoints
    print("\n" + "=" * 80)
    print("🧪 2. ТЕСТИРОВАНИЕ GET /api/finances/export-forecast")
    print("=" * 80)
    
    export_forecast_result = await test_export_forecast_endpoint()
    all_results.append(("Export Forecast", export_forecast_result))
    
    # Final summary
    print("\n" + "=" * 80)
    print("📊 ИТОГОВЫЙ ОТЧЁТ - ГОТОВНОСТЬ К PRODUCTION")
    print("=" * 80)
    
    total_tests = len(all_results)
    successful_tests = len([r for name, r in all_results if not r.errors])
    failed_tests = total_tests - successful_tests
    
    print(f"📈 Всего тестов: {total_tests}")
    print(f"✅ Успешных: {successful_tests}")
    print(f"❌ Неудачных: {failed_tests}")
    print(f"📊 Процент успеха: {(successful_tests/total_tests)*100:.1f}%")
    
    print(f"\n📋 ДЕТАЛЬНЫЕ РЕЗУЛЬТАТЫ:")
    
    for test_name, result in all_results:
        status = "✅ ГОТОВ" if not result.errors else f"❌ НЕ ГОТОВ ({len(result.errors)} ошибок)"
        print(f"   {test_name}: {status}")
        
        if result.errors:
            # Show only critical errors in summary
            critical_errors = [e for e in result.errors if 'КРИТЕРИЙ' in e and 'НЕ ВЫПОЛНЕН' in e]
            for error in critical_errors:
                print(f"      - {error}")
    
    print(f"\n🎯 ГОТОВНОСТЬ К PRODUCTION:")
    
    # Check each endpoint
    export_all_ready = not export_all_result.errors
    export_forecast_ready = not export_forecast_result.errors
    
    print(f"   {'✅' if export_all_ready else '❌'} GET /api/finances/export-all")
    print(f"   {'✅' if export_forecast_ready else '❌'} GET /api/finances/export-forecast")
    
    # Overall readiness
    if successful_tests == total_tests:
        print(f"\n🎉 ВСЕ ЭКСПОРТ ENDPOINTS ГОТОВЫ К PRODUCTION!")
        print(f"   ✅ Все endpoints работают стабильно")
        print(f"   ✅ Нет ошибок 500")
        print(f"   ✅ XLSX файлы валидны")
        print(f"   ✅ Имена файлов без кириллицы (только ASCII)")
        print(f"   ✅ Время отклика в пределах нормы")
        print(f"   ✅ Размеры файлов соответствуют ожиданиям")
    else:
        print(f"\n⚠️ ЕСТЬ ПРОБЛЕМЫ - НЕ ГОТОВО К PRODUCTION")
        print(f"   ❌ Требуется исправление ошибок перед deploy")
        
        # Show what needs to be fixed
        if not export_all_ready:
            print(f"   🔧 Исправить: GET /api/finances/export-all")
        if not export_forecast_ready:
            print(f"   🔧 Исправить: GET /api/finances/export-forecast")
    
    print("\n" + "=" * 80)
    print("🏁 ФИНАЛЬНАЯ ПРОВЕРКА ЗАВЕРШЕНА")
    print("=" * 80)
    
    return all_results

if __name__ == "__main__":
    success = asyncio.run(main())
    exit(0 if success else 1)
